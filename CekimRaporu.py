import streamlit as st
import streamlit.components.v1 as components
import requests
import json
import pandas as pd
from datetime import datetime, timedelta
import os
import pytz
import time
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO

# Sayfa konfigürasyonu
st.set_page_config(
    page_title="BetConstruct Raporlama Sistemi",
    page_icon="💰",
    layout="wide"
)

# Sağ üst köşeye yeşil nokta butonu ekle
st.markdown("""
<style>
.green-dot-container {
    position: fixed;
    top: 60px;
    right: 20px;
    z-index: 999;
}

.green-dot {
    width: 12px;
    height: 12px;
    background-color: #28a745;
    border-radius: 50%;
    cursor: pointer;
    box-shadow: 0 2px 4px rgba(0,0,0,0.2);
    transition: all 0.3s ease;
    animation: pulse-green 2s infinite;
}

.green-dot:hover {
    background-color: #1e7e34;
    transform: scale(1.2);
    box-shadow: 0 4px 8px rgba(0,0,0,0.3);
}

@keyframes pulse-green {
    0% {
        box-shadow: 0 0 0 0 rgba(40, 167, 69, 0.7);
    }
    70% {
        box-shadow: 0 0 0 10px rgba(40, 167, 69, 0);
    }
    100% {
        box-shadow: 0 0 0 0 rgba(40, 167, 69, 0);
    }
}

.tooltip {
    position: relative;
    display: inline-block;
}

.tooltip .tooltiptext {
    visibility: hidden;
    width: 120px;
    background-color: #333;
    color: white;
    text-align: center;
    border-radius: 6px;
    padding: 5px;
    position: absolute;
    z-index: 1;
    bottom: 125%;
    left: 50%;
    margin-left: -60px;
    opacity: 0;
    transition: opacity 0.3s;
    font-size: 12px;
}

.tooltip:hover .tooltiptext {
    visibility: visible;
    opacity: 1;
}
</style>

<div class="green-dot-container">
    <div class="tooltip">
        <a href="https://bonusraporu.streamlit.app/" target="_blank" style="text-decoration: none;">
            <div class="green-dot"></div>
        </a>
        <span class="tooltiptext">Bonus Raporu</span>
    </div>
</div>
""", unsafe_allow_html=True)

# Alternatif çözüm: Streamlit sidebar'da link butonu
with st.sidebar:
    st.markdown("---")
    if st.button("🟢 Bonus Raporu", help="Bonus Raporu uygulamasına git"):
        st.markdown("""
        <script>
        window.open('https://bonusraporu.streamlit.app/', '_blank');
        </script>
        """, unsafe_allow_html=True)
        st.info("🔗 Yeni sekmede açılıyor: https://bonusraporu.streamlit.app/")

# CSS stilleri ekleme
st.markdown("""
<style>
.status-approved {
    background-color: #d4edda;
    color: #155724;
    padding: 4px 8px;
    border-radius: 4px;
    font-weight: bold;
}

.status-rejected {
    background-color: #f8d7da;
    color: #721c24;
    padding: 4px 8px;
    border-radius: 4px;
    font-weight: bold;
}

.status-new {
    background-color: #cce5ff;
    color: #004085;
    padding: 4px 8px;
    border-radius: 4px;
    font-weight: bold;
}

.status-pending {
    background-color: #fff3cd;
    color: #856404;
    padding: 4px 8px;
    border-radius: 4px;
    font-weight: bold;
}

.action-buttons {
    position: fixed;
    top: 60px;
    right: 20px;
    z-index: 1000;
    background: white;
    padding: 15px;
    border-radius: 8px;
    box-shadow: 0 4px 20px rgba(0,0,0,0.2);
    border: 2px solid #007bff;
    min-width: 200px;
}

.approve-btn {
    background-color: #28a745;
    color: white;
    border: none;
    padding: 8px 16px;
    border-radius: 4px;
    cursor: pointer;
    margin: 2px;
    font-weight: bold;
}

.reject-btn {
    background-color: #dc3545;
    color: white;
    border: none;
    padding: 8px 16px;
    border-radius: 4px;
    cursor: pointer;
    margin: 2px;
    font-weight: bold;
}
</style>
""", unsafe_allow_html=True)
# Para formatlaması fonksiyonu
def format_turkish_currency(amount):
    """Türk Lirası formatında para birimi formatla (1.000,00 TL)"""
    try:
        if amount is None or amount == "" or (isinstance(amount, str) and amount.strip() == ""):
            return "0,00 TL"
        
        # String ise float'a çevir
        if isinstance(amount, str):
            amount = float(amount.replace(',', '.'))
        
        # Float'a çevir
        amount = float(amount)
        
        # Türk formatında formatla: 1.000,00 TL
        formatted = f"{amount:,.2f} TL"
        # Nokta ve virgülü değiştir (Türk formatı için)
        formatted = formatted.replace(',', 'X').replace('.', ',').replace('X', '.')
        
        return formatted
    
    except (ValueError, TypeError):
        return "0,00 TL"

# JavaScript tabanlı kopyalama butonu
def create_copy_button(text, button_text="📋 Kopyala", key=None):
    """JavaScript ile tek tıkla kopyalama butonu oluştur"""
    # Benzersiz ID oluştur
    button_id = f"copy_btn_{key}" if key else "copy_btn"
    text_id = f"copy_text_{key}" if key else "copy_text"
    
    # JavaScript kodu
    copy_script = f"""
    <div>
        <textarea id="{text_id}" style="width: 100%; height: 200px; font-family: monospace; font-size: 12px; padding: 10px; border: 1px solid #ddd; border-radius: 4px;" readonly>{text}</textarea>
        <br><br>
        <button id="{button_id}" onclick="copyToClipboard()" style="background-color: #ff4b4b; color: white; border: none; padding: 8px 16px; border-radius: 4px; cursor: pointer; font-weight: bold;">
            {button_text}
        </button>
        <span id="copy_status_{key}" style="margin-left: 10px; color: green; font-weight: bold;"></span>
    </div>
    
    <script>
    function copyToClipboard() {{
        const textArea = document.getElementById('{text_id}');
        const statusSpan = document.getElementById('copy_status_{key}');
        
        // Modern tarayıcılar için Clipboard API
        if (navigator.clipboard) {{
            navigator.clipboard.writeText(textArea.value).then(function() {{
                statusSpan.innerHTML = '✅ Kopyalandı!';
                setTimeout(() => statusSpan.innerHTML = '', 3000);
            }}).catch(function(err) {{
                // Fallback yöntemi
                fallbackCopy();
            }});
        }} else {{
            // Eski tarayıcılar için fallback
            fallbackCopy();
        }}
        
        function fallbackCopy() {{
            textArea.select();
            textArea.setSelectionRange(0, 99999);
            try {{
                document.execCommand('copy');
                statusSpan.innerHTML = '✅ Kopyalandı!';
                setTimeout(() => statusSpan.innerHTML = '', 3000);
            }} catch (err) {{
                statusSpan.innerHTML = '❌ Kopyalama başarısız - Manuel seçin';
                setTimeout(() => statusSpan.innerHTML = '', 5000);
            }}
        }}
    }}
    </script>
    """
    
    # HTML komponenti olarak render et
    components.html(copy_script, height=280)

# Token yönetimi fonksiyonları
def load_config():
    """Konfigürasyon dosyasından ayarları yükle"""
    config_file = "config.json"
    default_config = {
        "token": "2582007cbe97f891cf5fe69f4f2d44b002c021e6fca4c8276dc0accf4098d5fe",
        "api_url": "https://backofficewebadmin.betconstruct.com/api/tr/Client/GetClientWithdrawalRequestsWithTotals"
    }
    
    try:
        if os.path.exists(config_file):
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
                # Eksik alanları varsayılan değerlerle doldur
                for key, value in default_config.items():
                    if key not in config:
                        config[key] = value
                return config
        else:
            # İlk çalışma, varsayılan config dosyasını oluştur
            save_config(default_config)
            return default_config
    except Exception:
        return default_config

def save_config(config):
    """Konfigürasyonu dosyaya kaydet"""
    try:
        with open("config.json", 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        return True
    except Exception:
        return False

# Konfigürasyonu yükle
config = load_config()

def update_global_config():
    """Global değişkenleri güncelle"""
    global TOKEN, API_URL, config
    config = load_config()
    TOKEN = config.get("token", "")
    API_URL = config.get("api_url", "")

# İlk yükleme
update_global_config()

def get_headers():
    """Güncel token ile header bilgilerini al"""
    return {
        "Content-Type": "application/json;charset=UTF-8",
        "Accept": "application/json, text/plain, */*",
        "Authentication": TOKEN,
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
        "Origin": "https://backoffice.betconstruct.com",
        "Referer": "https://backoffice.betconstruct.com/"
    }

def fetch_client_details(client_id):
    """Müşteri detaylarını API'den getir"""
    url = f"https://backofficewebadmin.betconstruct.com/api/tr/Client/GetClientById?id={client_id}"
    headers = get_headers()
    
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            data = response.json()
            if not data.get("HasError", True):
                return data.get("Data")
            else:
                st.error(f"API Hatası: {data.get('AlertMessage', 'Bilinmeyen hata')}")
                return None
        else:
            st.error(f"HTTP Hatası: {response.status_code}")
            return None
    except Exception as e:
        st.error(f"Bağlantı hatası: {str(e)}")
        return None

def fetch_client_sportsbook_profile(client_id):
    """Müşteri spor bahis profilini getir"""
    url = f"https://backofficewebadmin.betconstruct.com/api/tr/Client/GetClientSportsBookProfiles"
    headers = get_headers()
    
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            data = response.json()
            if not data.get("HasError", True) and data.get("Data"):
                # Client ID'ye göre profili bul
                profiles = data.get("Data", [])
                for profile in profiles:
                    if str(profile.get("ClientId", "")) == str(client_id):
                        return profile
                return None
            else:
                return None
        else:
            return None
    except Exception as e:
        return None

def fetch_client_casino_profile(client_id):
    """Müşteri casino profilini getir"""
    url = f"https://backofficewebadmin.betconstruct.com/api/tr/Client/GetClientCasinoProfiles"
    headers = get_headers()
    
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            data = response.json()
            if not data.get("HasError", True) and data.get("Data"):
                # Client ID'ye göre profili bul
                profiles = data.get("Data", [])
                for profile in profiles:
                    if str(profile.get("ClientId", "")) == str(client_id):
                        return profile
                return None
            else:
                return None
        else:
            return None
    except Exception as e:
        return None

def fetch_client_kpis(client_id):
    """Müşteri KPI bilgilerini getir - yatırım/çekim tutarları ve sayıları"""
    url = "https://backofficewebadmin.betconstruct.com/api/tr/Client/GetClientKpis"
    headers = get_headers()
    headers['X-Requested-With'] = 'XMLHttpRequest'
    
    payload = {
        'ClientId': int(client_id)
    }
    
    try:
        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 200:
            data = response.json()
            if not data.get("HasError", True):
                return data.get("Data")
            else:
                st.error(f"KPI API Hatası: {data.get('AlertMessage', 'Bilinmeyen hata')}")
                return None
        else:
            st.error(f"KPI HTTP Hatası: {response.status_code}")
            return None
    except Exception as e:
        st.error(f"KPI Bağlantı hatası: {str(e)}")
        return None


# ==============================================
# BAHİS RAPORU İÇİN YENİ API FONKSİYONLARI
# ==============================================

def fetch_client_transactions(client_id, date_from=None, date_to=None, page=1, page_size=50):
    """Müşteri işlemlerini getir - bahis geçmişi için"""
    url = "https://backofficewebadmin.betconstruct.com/api/tr/Client/GetClientTransactionsV1"
    headers = get_headers()
    headers["X-Requested-With"] = "XMLHttpRequest"
    
    # Eğer tarih belirtilmemişse son 30 gün
    if not date_from:
        date_from = datetime.now() - timedelta(days=30)
    if not date_to:
        date_to = datetime.now()
    
    # Tarih formatını ayarla
    date_from_str = date_from.strftime("%Y-%m-%dT00:00:00.000Z")  
    date_to_str = date_to.strftime("%Y-%m-%dT23:59:59.999Z")
    
    payload = {
        "ClientId": int(client_id),
        "DateFrom": date_from_str,
        "DateTo": date_to_str,
        "Page": page,
        "PageSize": page_size
    }
    
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=30)
        if response.status_code == 200:
            data = response.json()
            if not data.get("HasError", True):
                return data.get("Data", {})
        return None
    except Exception as e:
        return None

def fetch_client_sports_bets(client_id, date_from=None, date_to=None):
    """Müşteri spor bahisleri getir"""
    url = "https://backofficewebadmin.betconstruct.com/api/tr/Client/GetClientSportBets"
    headers = get_headers()
    
    # Tarih formatını ayarla
    if not date_from:
        date_from = datetime.now() - timedelta(days=30)
    if not date_to:
        date_to = datetime.now()
        
    date_from_str = date_from.strftime("%Y-%m-%dT00:00:00.000Z")
    date_to_str = date_to.strftime("%Y-%m-%dT23:59:59.999Z")
    
    payload = {
        "ClientId": int(client_id),
        "DateFrom": date_from_str,
        "DateTo": date_to_str,
        "Page": 1,
        "PageSize": 100
    }
    
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=30)
        if response.status_code == 200:
            data = response.json()
            if not data.get("HasError", True):
                return data.get("Data", {})
        return None
    except:
        return None

def fetch_pending_sports_bets(client_id):
    """Bekleyen spor bahislerini getir"""
    url = "https://backofficewebadmin.betconstruct.com/api/tr/Client/GetClientPendingSportsBets"
    headers = get_headers()
    
    payload = {
        "ClientId": int(client_id)
    }
    
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=30)
        if response.status_code == 200:
            data = response.json()
            if not data.get("HasError", True):
                return data.get("Data", [])
        return []
    except:
        return []

def create_betting_report(client_id, client_transactions=None, sports_bets=None, pending_bets=None):
    """Bahis raporu oluştur"""
    try:
        # Eğer veriler yoksa API'den getir
        if not client_transactions:
            client_transactions = fetch_client_transactions(client_id)
        if not sports_bets:
            sports_bets = fetch_client_sports_bets(client_id)
        if not pending_bets:
            pending_bets = fetch_pending_sports_bets(client_id)
        
        report_sections = []
        
        # 1. Genel Bahis Raporu
        report_sections.append("📊 BAHİS RAPORU")
        report_sections.append("=" * 50)
        
        if client_transactions and "Transactions" in client_transactions:
            transactions = client_transactions["Transactions"]
            
            # Transaction türlerine göre grupla
            bets_data = []
            for transaction in transactions:
                if transaction.get("Type") in ["Bet", "SportBet", "CasinoBet"]:
                    game_name = transaction.get("GameName", "Bilinmeyen Oyun")
                    bet_amount = float(transaction.get("Amount", 0))
                    win_amount = float(transaction.get("WinAmount", 0))
                    
                    bets_data.append({
                        "game": game_name,
                        "bet": bet_amount,
                        "win": win_amount
                    })
            
            if bets_data:
                report_sections.append("Oyun Adı                    Bahis Miktarı       Kazanç")
                report_sections.append("-" * 60)
                
                for bet in bets_data[:20]:  # İlk 20 bahis
                    game_name = bet["game"][:25].ljust(25)
                    bet_amount = f"{bet['bet']:.2f} TL".rjust(12)
                    win_amount = f"{bet['win']:.2f} TL".rjust(12)
                    report_sections.append(f"{game_name} {bet_amount} {win_amount}")
                
                # 2. Tek Seferdeki En Yüksek Bahis
                highest_bet = max(bets_data, key=lambda x: x["bet"])
                report_sections.append("")
                report_sections.append("🎯 TEK SEFERDEKİ EN YÜKSEK BAHİS")
                report_sections.append("-" * 50)
                report_sections.append(f"Oyun Adı: {highest_bet['game']}")
                report_sections.append(f"Bahis Miktarı: {highest_bet['bet']:.2f} TL")
                report_sections.append(f"Kazanç: {highest_bet['win']:.2f} TL")
                
                # 3. Tek Seferde En Yüksek Kazanç
                highest_win = max(bets_data, key=lambda x: x["win"])
                report_sections.append("")
                report_sections.append("💰 TEK SEFERDE EN YÜKSEK KAZANÇ")
                report_sections.append("-" * 50)
                report_sections.append(f"Oyun Adı: {highest_win['game']}")
                report_sections.append(f"Bahis Miktarı: {highest_win['bet']:.2f} TL")
                report_sections.append(f"Kazanç: {highest_win['win']:.2f} TL")
            else:
                report_sections.append("⚠️ Bahis işlemi bulunamadı")
        
        # 4. Bekleyen Spor Bahisleri
        report_sections.append("")
        report_sections.append("⏳ BEKLEYEN SPOR BAHİSLERİ")
        report_sections.append("-" * 50)
        
        if pending_bets and len(pending_bets) > 0:
            report_sections.append("Bahis Miktarı          Oran")
            report_sections.append("-" * 30)
            
            for bet in pending_bets[:10]:  # İlk 10 bekleyen bahis
                bet_amount = f"{float(bet.get('Amount', 0)):.2f} TL".ljust(15)
                odds = f"{float(bet.get('Odds', 0)):.2f}".rjust(8)
                report_sections.append(f"{bet_amount} {odds}")
        else:
            report_sections.append("✅ Bekleyen spor bahsi yok")
        
        return "\n".join(report_sections)
        
    except Exception as e:
        return f"❌ Bahis raporu oluşturma hatası: {str(e)}"


# Login verilerini API'den getir

def fetch_client_logins(client_id):
    """Müşteri login verilerini GetLogins API'sinden getir"""
    url = "https://backofficewebadmin.betconstruct.com/api/tr/Client/GetLogins"
    headers = get_headers()
    
    try:
        payload = {
            "Page": 1,
            "PageSize": 100,
            "ClientId": client_id
        }
        
        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 200:
            data = response.json()
            if not data.get("HasError", True) and data.get("Data"):
                return data["Data"]["Objects"]
            else:
                st.error(f"Login verileri alınamadı: {data.get('AlertMessage', 'Bilinmeyen hata')}")
                return None
        else:
            st.error(f"HTTP Hatası: {response.status_code}")
            return None
    except Exception as e:
        st.error(f"Login verileri alınamadı: {str(e)}")
        return None

# Durum mapping - API'den gelen gerçek durumlar
STATUS_MAPPING = {
    "Tümü": [],
    "Yeni": ["Yeni", "New"],
    "İzin Verildi": ["İzin Verildi", "Approved"], 
    "Beklemede": ["Beklemede", "Pending"],
    "Reddedildi": ["Reddedildi", "Rejected"],
    "İptal edildi": ["İptal edildi", "Cancelled"],
    "Ödendi": ["Ödendi", "Paid"],
    "RollBacked": ["RollBacked", "Geri Alındı"]
}

def get_colored_status(status_text):
    """Durum metnini renkli HTML olarak döndür"""
    if status_text in ["Ödendi", "Paid", "Completed", "Processed"]:
        return f'<span class="status-approved">✅ {status_text}</span>'
    elif status_text in ["Reddedildi", "Rejected", "Declined"]:
        return f'<span class="status-rejected">❌ {status_text}</span>'
    elif status_text in ["Yeni", "New"]:
        return f'<span class="status-new">🆕 {status_text}</span>'
    else:
        return f'<span class="status-pending">⚠️ {status_text}</span>'

def approve_withdrawal_request(request_id, client_id):
    """Çekim talebini onayla"""
    try:
        approve_url = "https://backofficewebadmin.betconstruct.com/api/tr/Client/ApproveWithdrawalRequest"
        
        payload = {
            "RequestId": request_id,
            "ClientId": client_id
        }
        
        response = requests.post(approve_url, json=payload, headers=get_headers())
        
        if response.status_code == 200:
            return True, "Talep başarıyla onaylandı"
        else:
            return False, f"Onay hatası: {response.status_code}"
    except Exception as e:
        return False, f"Onay hatası: {str(e)}"

def reject_withdrawal_request(request_id, client_id, reason="Fraud analizi sonucu reddedildi"):
    """Çekim talebini reddet"""
    try:
        reject_url = "https://backofficewebadmin.betconstruct.com/api/tr/Client/RejectWithdrawalRequest"
        
        payload = {
            "RequestId": request_id,
            "ClientId": client_id,
            "Reason": reason
        }
        
        response = requests.post(reject_url, json=payload, headers=get_headers())
        
        if response.status_code == 200:
            return True, "Talep başarıyla reddedildi"
        else:
            return False, f"Red hatası: {response.status_code}"
    except Exception as e:
        return False, f"Red hatası: {str(e)}"

def parse_api_datetime(date_str):
    """API'den gelen tarih string'ini datetime objesine çevir"""
    try:
        if not date_str:
            return None
        
        # ISO 8601 formatı with timezone (+04:00 gibi)
        if '+' in date_str:
            # 2025-07-16T09:20:19.535+04:00 formatı
            dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            return dt
        elif 'T' in date_str:
            # 2025-07-16T08:20:19.535 formatı (timezone yok)
            dt = datetime.fromisoformat(date_str)
            return dt
        else:
            return None
    except Exception as e:
        st.error(f"Tarih parse hatası: {date_str} - {str(e)}")
        return None

def is_date_in_range(request_date, date_from, date_to):
    """Bir tarihin belirtilen aralıkta olup olmadığını kontrol et"""
    try:
        if not request_date:
            return False
        
        # Timezone bilgisini kaldır ve sadece tarih kıyaslaması yap
        request_date_naive = request_date.replace(tzinfo=None)
        
        # Gün bazında karşılaştırma
        request_day = request_date_naive.date()
        
        return date_from <= request_day <= date_to
    except Exception as e:
        st.error(f"Tarih karşılaştırma hatası: {str(e)}")
        return False

def filter_requests_by_date(requests, date_from, date_to):
    """İstemci tarafında tarih filtrelemesi yap"""
    filtered_requests = []
    
    for request in requests:
        if not isinstance(request, dict):
            continue
        
        # RequestTime ve RequestTimeLocal alanlarını kontrol et
        request_time = parse_api_datetime(request.get("RequestTime"))
        request_time_local = parse_api_datetime(request.get("RequestTimeLocal"))
        
        # Her iki tarih alanını da kontrol et
        date_matches = False
        
        if request_time and is_date_in_range(request_time, date_from, date_to):
            date_matches = True
        elif request_time_local and is_date_in_range(request_time_local, date_from, date_to):
            date_matches = True
        
        if date_matches:
            filtered_requests.append(request)
    
    return filtered_requests

def fetch_withdrawal_requests(date_from, date_to, statuses, page=1, page_size=100, timezone_format="UTC"):
    """API'den çekim taleplerini getir"""
    # BetConstruct API formatı - farklı timezone formatları test et
    
    # Timezone formatına göre tarih formatını belirle
    if timezone_format == "TR_PLUS3":
        date_from_str = date_from.strftime("%Y-%m-%dT00:00:00+03:00")
        date_to_str = date_to.strftime("%Y-%m-%dT23:59:59+03:00")
    elif timezone_format == "TR_PLUS4":
        date_from_str = date_from.strftime("%Y-%m-%dT00:00:00+04:00")
        date_to_str = date_to.strftime("%Y-%m-%dT23:59:59+04:00")
    elif timezone_format == "UTC_Z":
        date_from_str = date_from.strftime("%Y-%m-%dT00:00:00.000Z")
        date_to_str = date_to.strftime("%Y-%m-%dT23:59:59.999Z")
    elif timezone_format == "LOCAL_TR":
        # Türkiye yerel saati (UTC+3 kış, UTC+4 yaz)
        import pytz
        tr_tz = pytz.timezone('Europe/Istanbul')
        dt_from = tr_tz.localize(datetime.combine(date_from, datetime.min.time()))
        dt_to = tr_tz.localize(datetime.combine(date_to, datetime.max.time()))
        date_from_str = dt_from.isoformat()
        date_to_str = dt_to.isoformat()
    else:  # UTC default
        date_from_str = date_from.strftime("%Y-%m-%dT00:00:00")
        date_to_str = date_to.strftime("%Y-%m-%dT23:59:59")
    
    payload = {
        "DateFrom": date_from_str,
        "DateTo": date_to_str,
        "PaymentMethodId": None,
        "Statuses": statuses if statuses else [],
        "Page": page,
        "PageSize": page_size,
        "ClientId": None,
        "ClientName": None,
        "ClientUsername": None,
        "PaymentSystemName": None
    }
    
    try:
        headers = get_headers()
        response = requests.post(API_URL, headers=headers, json=payload, timeout=30)
        if response.status_code == 200:
            result = response.json()
            return result
        else:
            st.error(f"API Hatası: {response.status_code}")
            st.error(f"Yanıt: {response.text}")
            return None
    except Exception as e:
        st.error(f"Bağlantı hatası: {str(e)}")
        return None

def parse_bank_info(info_text):
    """Banka bilgilerini parse et - BANKA HAVALE formatından"""
    if not info_text:
        return {"isim": "", "iban": "", "banka": ""}
    
    try:
        # JSON formatı kontrolü
        if isinstance(info_text, str) and info_text.startswith('{'):
            info = json.loads(info_text)
            return {
                "isim": info.get("AccountHolderName", ""),
                "iban": info.get("IBAN", ""),
                "banka": info.get("BankName", "")
            }
        
        # BANKA HAVALE formatı parse et
        elif isinstance(info_text, str) and "BANKA HAVALE" in info_text:
            isim = ""
            iban = ""
            banka = ""
            
            # Hesap Adı ve Soyadı
            if "Hesap Adi ve Soyadi:" in info_text:
                start = info_text.find("Hesap Adi ve Soyadi:") + len("Hesap Adi ve Soyadi:")
                end = info_text.find(",", start)
                if end == -1:
                    end = len(info_text)
                isim = info_text[start:end].strip()
            
            # Banka Adı
            if "Banka Adi:" in info_text:
                start = info_text.find("Banka Adi:") + len("Banka Adi:")
                end = info_text.find(",", start)
                if end == -1:
                    end = len(info_text)
                banka = info_text[start:end].strip()
            
            # IBAN Numarası
            if "IBAN Numarasi:" in info_text:
                start = info_text.find("IBAN Numarasi:") + len("IBAN Numarasi:")
                end = info_text.find(",", start)
                if end == -1:
                    end = len(info_text)
                iban = info_text[start:end].strip()
            
            return {"isim": isim, "iban": iban, "banka": banka}
        
        # Diğer formatlar için boş döndür
        else:
            return {"isim": "", "iban": "", "banka": ""}
            
    except Exception as e:
        return {"isim": "", "iban": "", "banka": ""}

def create_withdrawal_report(selected_requests):
    """Seçilen talepler için çekim raporu oluştur"""
    if not selected_requests:
        return ""
    
    report = ""
    for request in selected_requests:
        if request.get("PaymentSystemName") == "BankTransferBME":
            bank_info = parse_bank_info(request.get("Info", ""))
            formatted_amount = format_turkish_currency(request.get('Amount', 0))
            
            report += f"İsimSoyisim : {bank_info['isim']}\n"
            report += f"İban : {bank_info['iban']}\n"
            report += f"Banka : {bank_info['banka']}\n"
            report += f"Miktar : {formatted_amount}\n"
            report += "-" * 40 + "\n"
    
    return report

def create_fraud_report(withdrawal_request, client_id):
    """Fraud raporu oluştur"""
    try:
        # Withdrawal request'den temel bilgiler
        request_amount = withdrawal_request.get("Amount", 0)
        payment_method = withdrawal_request.get("PaymentSystemName", "-")
        full_name = withdrawal_request.get("ClientName", "")
        username = withdrawal_request.get("ClientLogin", "")
        
        # Client detaylarını API'den getir
        client_details = fetch_client_details(client_id)
        client_kpis = fetch_client_kpis(client_id)
        
        # Login verilerini API'den getir
        login_data = fetch_client_logins(client_id)
        
        # Varsayılan değerler
        current_balance = 0
        game_type = "-"
        last_deposit = 0
        total_deposits = 0
        total_withdrawals = 0
        withdrawal_count = 0
        deposit_count = 0
        game_status = "Bilinmiyor"
        game_desc = ""
        
        if client_details and client_kpis and login_data:
            # İlk olarak KPI verilerini al
            kpi_data = client_kpis[0]
            
            # Güncel API verilerine göre bilgileri al
            current_balance = float(client_details.get("Balance", 0))
            
            # Son aktivite bilgileri
            last_sport_bet = kpi_data.get("LastSportBetTime", "")
            last_casino_bet = kpi_data.get("LastCasinoBetTime", "")
            last_login = client_details.get("LastLoginLocalDate", "")
            
            # Oyun türünü belirle ve detaylı analiz yap
            try:
                if last_casino_bet:
                    casino_active = (datetime.now() - datetime.strptime(last_casino_bet.split('+')[0], '%Y-%m-%dT%H:%M:%S.%f')).days < 30
                else:
                    casino_active = False
            except Exception as e:
                st.error(f"Error parsing casino bet time '{last_casino_bet}': {str(e)}")
                casino_active = False
            
            try:
                if last_sport_bet:
                    sport_active = (datetime.now() - datetime.strptime(last_sport_bet.split('+')[0], '%Y-%m-%dT%H:%M:%S.%f')).days < 30
                else:
                    sport_active = False
            except Exception as e:
                st.error(f"Error parsing sport bet time '{last_sport_bet}': {str(e)}")
                sport_active = False
            
            # Aktiflik süresi hesapla - YENİ DÜZENLENMİŞ HALİ
            active_days = 0  # Başlangıç değeri
            game_status = "Aktivite bilgisi alınamadı"
            game_desc = "- Aktivite bilgisi alınamadı\n"
            
            # Tarih parse fonksiyonu
            def parse_date(date_str):
                try:
                    # Önce UTC formatında parse et
                    date = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                    # Sistem zaman dilimine çevir
                    date = date.astimezone(pytz.timezone('Turkey'))
                    return date
                except Exception as e:
                    st.error(f"Tarih parse hatası: {str(e)}")
                    return None
            
            # Sistem zaman diliminde mevcut zaman
            now = datetime.now(pytz.timezone('Turkey'))
            
            # Login tarihini parse et
            login_time = None
            if last_login:
                login_time = parse_date(last_login)
            
            # Kayıt tarihini parse et
            reg_date = None
            reg_date_str = client_details.get("RegistrationDate")
            if reg_date_str:
                reg_date = parse_date(reg_date_str)
            
            # Fallback stratejisi
            if login_time and reg_date:
                # Her iki tarih de mevcutsa, kayıt tarihinden son girişe kadar olan süreyi al
                total_days = (login_time - reg_date).days
                months_passed = total_days / 30
                avg_active_per_month = total_days / months_passed if months_passed > 0 else 0
                active_days = total_days
                game_desc = f"- Aktiflik süresi: {active_days} gün\n"
                game_desc += f"- Ayda ortalama aktif gün: {avg_active_per_month:.1f} gün\n"
                game_status = f"{active_days} gün aktif"
            elif login_time:
                # Sadece login tarihi mevcutsa, bugünden login tarihine kadar olan süreyi al
                active_days = (now - login_time).days
                game_desc = f"- Aktiflik süresi: {active_days} gün\n"
                game_status = f"{active_days} gün aktif"
            elif reg_date:
                # Sadece kayıt tarihi mevcutsa, bugünden kayıt tarihine kadar olan süreyi al
                active_days = (now - reg_date).days
                game_desc = f"- Aktiflik süresi: {active_days} gün\n"
                game_status = f"{active_days} gün aktif"
            elif login_data:
                try:
                    login_dates = []
                    for login in login_data:
                        if login.get("StartTime"):
                            try:
                                login_time = parse_date(login["StartTime"])
                                if login_time:
                                    login_dates.append(login_time)
                            except Exception as e:
                                continue
                    
                    if login_dates:
                        first_login = min(login_dates)
                        active_days = (now - first_login).days
                        game_desc = f"- Aktiflik süresi: {active_days} gün\n"
                        game_status = f"{active_days} gün aktif"
                except Exception as e:
                    st.error(f"Login veri analizi hatası: {str(e)}")
            else:
                game_status = "Hiç giriş yapmamış"
                game_desc = "- Kayıtlı giriş bulunamadı\n"
                
            # Günlük oyun süresi hesapla (login verilerinden)
            if login_data:
                # Session verilerini hesapla
                valid_sessions = []
                daily_playtimes = {}
                total_play_time = 0
                days_with_activity = 0
                activity_percentage = 0
                longest_session = {'duration': 0, 'start_time': datetime.now()}
                shortest_session = {'duration': 0, 'start_time': datetime.now()}
                
                for login in login_data:
                    if login.get("EndTime") and login.get("StartTime"):
                        try:
                            # Remove timezone info and try both formats
                            end_time_str = login["EndTime"].split('+')[0]
                            start_time_str = login["StartTime"].split('+')[0]
                            
                            # Try both formats (with and without microseconds)
                            try:
                                end_time = datetime.strptime(end_time_str, '%Y-%m-%dT%H:%M:%S.%f')
                            except ValueError:
                                end_time = datetime.strptime(end_time_str, '%Y-%m-%dT%H:%M:%S')
                            
                            try:
                                start_time = datetime.strptime(start_time_str, '%Y-%m-%dT%H:%M:%S.%f')
                            except ValueError:
                                start_time = datetime.strptime(start_time_str, '%Y-%m-%dT%H:%M:%S')
                            
                            duration = (end_time - start_time).total_seconds() / 3600  # Saat cinsinden
                            total_play_time += duration
                            
                            # Günlük toplam oyun süresini hesapla
                            day = start_time.date()
                            daily_playtimes[day] = daily_playtimes.get(day, 0) + duration
                            
                            # Session verilerini topla
                            valid_sessions.append({
                                "duration": duration,
                                "day": day,
                                "start_time": start_time,
                                "end_time": end_time
                            })
                        except Exception as e:
                            continue
                
                # Ortalama günlük oyun süresi
                avg_daily_play = total_play_time / len(daily_playtimes) if daily_playtimes else 0
                
                # Ortalama oturum süresi
                avg_session_duration = sum(s["duration"] for s in valid_sessions) / len(valid_sessions) if valid_sessions else 0
                
                # Günlük aktivite bilgilerini hesapla
                if valid_sessions:
                    longest_session = max(valid_sessions, key=lambda x: x["duration"])
                    shortest_session = min(valid_sessions, key=lambda x: x["duration"])
                    days_with_activity = len(daily_playtimes)
                    activity_percentage = (days_with_activity / active_days) * 100 if active_days > 0 else 0
                
                # Analiz sonuçlarını rapora ekle
                game_desc += "\n=== ZAMAN ANALİZİ ===\n"
                game_desc += f"- Toplam aktif gün: {days_with_activity}/{active_days} (%{int(activity_percentage)})\n"
                game_desc += f"- Ortalama günlük oyun süresi: {avg_daily_play:.1f} saat\n"
                game_desc += f"- Ortalama oturum süresi: {avg_session_duration:.1f} saat\n"
                game_desc += f"- En uzun oturum: {longest_session['duration']:.1f} saat ({longest_session['start_time'].strftime('%d.%m.%Y')})\n"
                game_desc += f"- En kısa oturum: {shortest_session['duration']:.1f} saat ({shortest_session['start_time'].strftime('%d.%m.%Y')})\n"
            
            # Son 30 gün içindeki farklı IP kullanımı
            unique_ips = set()
            if login_data:
                thirty_days_ago = datetime.now() - timedelta(days=30)
                for login in login_data:
                    if login.get("StartTime"):
                        try:
                            login_time = datetime.strptime(login["StartTime"].split('+')[0], '%Y-%m-%dT%H:%M:%S.%f')
                        except ValueError:
                            try:
                                login_time = datetime.strptime(login["StartTime"].split('+')[0], '%Y-%m-%dT%H:%M:%S')
                            except Exception as e:
                                st.error(f"Error parsing login time '{login['StartTime']}': {str(e)}")
                                continue
                        
                        if login_time >= thirty_days_ago:
                            unique_ips.add(login.get("LoginIP", ""))
            ip_changes = len(unique_ips)
            
            # KPI verilerinden yatırım/çekim bilgilerini al
            total_deposits = float(kpi_data.get("TotalDeposit", 0))
            total_withdrawals = float(kpi_data.get("TotalWithdrawal", 0))
            withdrawal_count = int(kpi_data.get("WithdrawalCount", 0))
            deposit_count = int(kpi_data.get("DepositCount", 0))
            last_deposit = float(kpi_data.get('LastDepositAmount', 0))            
            if casino_active and sport_active:
                game_type = "Casino & Spor"
                game_desc = "- Hem casino hem spor bahisleri oynuyor\n"
                game_desc += f"- Ortalama günlük oyun süresi: {avg_daily_play:.1f} saat\n"
                game_desc += f"- Son 30 günde {ip_changes} farklı IP kullanımı\n"
                
                # Casino detayları
                if kpi_data.get("TotalCasinoStakes", 0) > 0:
                    casino_ratio = float(kpi_data.get("TotalCasinoStakes", 0)) / (float(kpi_data.get("TotalSportStakes", 0)) + float(kpi_data.get("TotalCasinoStakes", 0)))
                    game_desc += f"- Casino oyun oranı: %{int(casino_ratio * 100)}\n"
                    
                    # Spesifik casino tercihleri
                    if kpi_data.get("TotalSlotStakes", 0) > 0:
                        slot_ratio = float(kpi_data.get("TotalSlotStakes", 0)) / float(kpi_data.get("TotalCasinoStakes", 0))
                        game_desc += f"- Slot oyun oranı: %{int(slot_ratio * 100)}\n"
                    if kpi_data.get("TotalLiveCasinoStakes", 0) > 0:
                        live_ratio = float(kpi_data.get("TotalLiveCasinoStakes", 0)) / float(kpi_data.get("TotalCasinoStakes", 0))
                        game_desc += f"- Live Casino oranı: %{int(live_ratio * 100)}\n"
                
                # Spor detayları
                if kpi_data.get("TotalSportStakes", 0) > 0:
                    sport_ratio = float(kpi_data.get("TotalSportStakes", 0)) / (float(kpi_data.get("TotalSportStakes", 0)) + float(kpi_data.get("TotalCasinoStakes", 0)))
                    game_desc += f"- Spor bahis oranı: %{int(sport_ratio * 100)}\n"
                    
                    # Spor türü analizi
                    if kpi_data.get("TotalFootballStakes", 0) > 0:
                        football_ratio = float(kpi_data.get("TotalFootballStakes", 0)) / float(kpi_data.get("TotalSportStakes", 0))
                        game_desc += f"- Futbol bahis oranı: %{int(football_ratio * 100)}\n"
                    if kpi_data.get("TotalBasketballStakes", 0) > 0:
                        basket_ratio = float(kpi_data.get("TotalBasketballStakes", 0)) / float(kpi_data.get("TotalSportStakes", 0))
                        game_desc += f"- Basketbol bahis oranı: %{int(basket_ratio * 100)}\n"
            
            elif casino_active:
                game_type = "Casino"
                game_desc = "- Ağırlıklı casino oyuncusu\n"
                game_desc += f"- Ortalama günlük oyun süresi: {avg_daily_play:.1f} saat\n"
                game_desc += f"- Son 30 günde {ip_changes} farklı IP kullanımı\n"
                
                # Casino detayları
                if kpi_data.get("TotalCasinoStakes", 0) > 0:
                    game_desc += f"- Casino oyun oranı: %{int(100)}\n"
                    
                    # Slot/Live Casino tercihini bul
                    if kpi_data.get("TotalSlotStakes", 0) > kpi_data.get("TotalLiveCasinoStakes", 0):
                        game_desc += "- Ağırlıklı Slot oyunları\n"
                        if kpi_data.get("TotalSlotStakes", 0) > 0:
                            slot_ratio = float(kpi_data.get("TotalSlotStakes", 0)) / float(kpi_data.get("TotalCasinoStakes", 0))
                            game_desc += f"- Slot oyun oranı: %{int(slot_ratio * 100)}\n"
                    else:
                        game_desc += "- Ağırlıklı Live Casino\n"
                        if kpi_data.get("TotalLiveCasinoStakes", 0) > 0:
                            live_ratio = float(kpi_data.get("TotalLiveCasinoStakes", 0)) / float(kpi_data.get("TotalCasinoStakes", 0))
                            game_desc += f"- Live Casino oranı: %{int(live_ratio * 100)}\n"
            
            elif sport_active:
                game_type = "Spor Bahis"
                game_desc = "- Ağırlıklı spor bahisleri oynuyor\n"
                game_desc += f"- Ortalama günlük oyun süresi: {avg_daily_play:.1f} saat\n"
                game_desc += f"- Son 30 günde {ip_changes} farklı IP kullanımı\n"
                
                # Spor detayları
                if kpi_data.get("TotalSportStakes", 0) > 0:
                    game_desc += f"- Spor bahis oranı: %{int(100)}\n"
                    
                    # Spor türü analizi
                    if kpi_data.get("TotalFootballStakes", 0) > kpi_data.get("TotalBasketballStakes", 0):
                        game_desc += "- Ağırlıklı Futbol bahisleri\n"
                        if kpi_data.get("TotalFootballStakes", 0) > 0:
                            football_ratio = float(kpi_data.get("TotalFootballStakes", 0)) / float(kpi_data.get("TotalSportStakes", 0))
                            game_desc += f"- Futbol bahis oranı: %{int(football_ratio * 100)}\n"
                    else:
                        game_desc += "- Ağırlıklı Basketbol bahisleri\n"
                        if kpi_data.get("TotalBasketballStakes", 0) > 0:
                            basket_ratio = float(kpi_data.get("TotalBasketballStakes", 0)) / float(kpi_data.get("TotalSportStakes", 0))
                            game_desc += f"- Basketbol bahis oranı: %{int(basket_ratio * 100)}\n"
            
            else:
                game_type = "Pasif"
                game_desc = "- Son 30 günde aktivite yok\n"
                
            # Finansal analiz ekle
            if total_deposits > 0 and total_withdrawals > 0:
                deposit_withdrawal_ratio = total_withdrawals / total_deposits
                avg_deposit = total_deposits/deposit_count if deposit_count > 0 else 0
                avg_withdrawal = total_withdrawals/withdrawal_count if withdrawal_count > 0 else 0
                
                game_desc += f"- Yatırım/Çekim oranı: %{int(deposit_withdrawal_ratio * 100)}\n"
                game_desc += f"- Ortalama yatırım: {format_turkish_currency(avg_deposit)}\n"
                game_desc += f"- Ortalama çekim: {format_turkish_currency(avg_withdrawal)}\n"
                
                # Yatırım/çekim trend analizi
                if total_deposits > total_withdrawals:
                    game_desc += "- Yatırım odaklı profil\n"
                else:
                    game_desc += "- Çekim odaklı profil\n"
            
            # Davranış analizi ekle
            if login_data:
                # Login pattern analizi
                login_times = []
                for login in login_data:
                    try:
                        login_time = datetime.strptime(login["StartTime"].split('+')[0], '%Y-%m-%dT%H:%M:%S.%f')
                    except ValueError:
                        try:
                            login_time = datetime.strptime(login["StartTime"].split('+')[0], '%Y-%m-%dT%H:%M:%S')
                        except Exception as e:
                            st.error(f"Error parsing login time '{login['StartTime']}': {str(e)}")
                            continue
                    login_times.append(login_time)
                
                login_hours = [login_time.hour for login_time in login_times]
                
                # En yoğun giriş saatleri
                hour_counts = {}
                for hour in login_hours:
                    hour_counts[hour] = hour_counts.get(hour, 0) + 1
                most_active_hours = sorted(hour_counts.items(), key=lambda x: x[1], reverse=True)[:3]
                
                # Cihaz kaynakları analizi
                device_sources = {}
                for login in login_data:
                    source = login.get("SourceName", "Bilinmiyor")
                    device_sources[source] = device_sources.get(source, 0) + 1
                
                # Session duration distribution
                session_durations = []
                for login in login_data:
                    if login.get("EndTime") and login.get("StartTime"):
                        # Remove timezone info and try both formats
                        end_time_str = login["EndTime"].split('+')[0]
                        start_time_str = login["StartTime"].split('+')[0]
                        
                        # Try both formats (with and without microseconds)
                        end_time = parse_api_datetime(end_time_str)
                        start_time = parse_api_datetime(start_time_str)
                        duration = (end_time - start_time).total_seconds() / 3600  # Saat cinsinden
                        session_durations.append(duration)
                
                # Geographical IP analysis
                ip_countries = {
                    "178.241.183.24": "Türkiye",
                    "178.243.86.35": "Türkiye",
                    "178.243.75.100": "Türkiye",
                    "176.90.140.122": "Türkiye"
                }
                
                # Time-based activity patterns
                daily_active_hours = {}
                for login in login_data:
                    login_time = parse_api_datetime(login["StartTime"])
                    if login_time:
                        hour = login_time.hour
                        daily_active_hours[hour] = daily_active_hours.get(hour, 0) + 1
                game_desc += f"- En yoğun giriş saatleri: {', '.join([f'{h}:00 ({c} kez)' for h, c in most_active_hours])}\n"
                game_desc += f"- En çok kullanılan cihaz: {max(device_sources.items(), key=lambda x: x[1])[0]}\n"
                
                if session_durations:
                    avg_duration = sum(session_durations) / len(session_durations)
                    game_desc += f"- Ortalama oturum süresi: {avg_duration:.1f} saat\n"
                game_desc += f"- Aktiflik süresi: {active_days} gün\n"
                
                # Zaman dilimi analizi
                time_periods = {
                    (0, 6): "Gece",
                    (6, 12): "Sabah",
                    (12, 18): "Öğleden sonra",
                    (18, 24): "Akşam"
                }
                
                period_counts = {period: 0 for period in time_periods.values()}
                for hour in login_hours:
                    for (start, end), period in time_periods.items():
                        if start <= hour < end:
                            period_counts[period] += 1
                            break
                
                most_active_period = max(period_counts.items(), key=lambda x: x[1])[0]
                game_desc += f"- En çok aktif zaman dilimi: {most_active_period}\n"
            
            # KPI verilerinden yatırım/çekim bilgilerini al
            kpi_data = client_kpis[0]
            
            total_deposits = float(kpi_data.get("TotalDeposit", 0))
            total_withdrawals = float(kpi_data.get("TotalWithdrawal", 0))
            withdrawal_count = int(kpi_data.get("WithdrawalCount", 0))
            deposit_count = int(kpi_data.get("DepositCount", 0))
            last_deposit = float(kpi_data.get('LastDepositAmount', 0))
            
            # Activity durumu
            if last_login:
                login_time = parse_api_datetime(last_login)
                if login_time:
                    now = datetime.now()
                    days_diff = (now - login_time).days
                    if days_diff == 0:
                        game_status = "Evet"
                    elif days_diff <= 3:
                        game_status = "Evet"
                    else:
                        game_status = "Hayır"
                else:
                    game_status = "Bilinmiyor"
            else:
                game_status = "Hayır"
            
        else:
            st.warning("⚠️ Client detayları alınamadı - varsayılan değerler kullanılıyor")
        
        # Para değerlerini formatla
        formatted_request_amount = format_turkish_currency(request_amount)
        formatted_last_deposit = format_turkish_currency(last_deposit)
        formatted_current_balance = format_turkish_currency(current_balance)
        formatted_total_deposits = format_turkish_currency(total_deposits)
        formatted_total_withdrawals = format_turkish_currency(total_withdrawals)
        
        # Fraud raporu formatı
        report = f"""İsim Soyisim   : {full_name}
K. Adı         : {username}
Talep Miktarı  : {formatted_request_amount}
Talep yöntemi  : {payment_method}
Yatırım Miktarı: {formatted_last_deposit}
Oyun Türü      : {game_type}
Arka Bakiye    : {formatted_current_balance}
Oyuna Devam    : {game_status}

T. Yatırım Miktarı: {formatted_total_deposits}
T. Çekim Miktarı  : {formatted_total_withdrawals}
T. Çekim Adedi    : {withdrawal_count}
T. Yatırım Adedi  : {deposit_count}
Açıklama          : 

{game_desc}"""
        
        return report
        
    except Exception as e:
        st.error(f"Fraud raporu oluşturma hatası: {str(e)}")
        return None

def check_new_requests():
    """Yeni talepleri kontrol et"""
    try:
        # Bugünün tarihini al
        today = datetime.now().date()
        
        # API'den son talepleri getir
        result = fetch_withdrawal_requests(
            datetime.combine(today, datetime.min.time()),
            datetime.combine(today, datetime.max.time()),
            [],
            page_size=50,
            timezone_format="UTC"
        )
        
        if result and isinstance(result, dict) and "Data" in result:
            data_section = result["Data"]
            if "ClientRequests" in data_section:
                requests_data = data_section["ClientRequests"]
                
                # Sadece "Yeni" durumundaki talepleri filtrele
                new_requests = [req for req in requests_data 
                              if isinstance(req, dict) and req.get("StateName") in ["Yeni", "New"]]
                
                return new_requests
    except Exception as e:
        st.error(f"Yeni talep kontrolü hatası: {str(e)}")
        return []
    
    return []

def sort_requests_by_status_and_date(requests):
    """Talepleri duruma ve tarihe göre sırala - Yeni talepler en üstte"""
    if not requests:
        return []
    
    def sort_key(request):
        if not isinstance(request, dict):
            return (99, datetime.min)
        
        # Durum önceliği - Yeni talepler en üstte
        status = request.get("StateName", "")
        if status in ["Yeni", "New"]:
            priority = 0
        elif status in ["Beklemede", "Pending"]:
            priority = 1
        elif status in ["İzin Verildi", "Approved"]:
            priority = 2
        else:
            priority = 3
        
        # Tarih - En yeni en üstte
        try:
            request_time = parse_api_datetime(request.get("RequestTime"))
            if request_time:
                return (priority, -request_time.timestamp())  # Negatif ile tersten sırala
        except:
            pass
        
        return (priority, 0)
    
    return sorted(requests, key=sort_key)

# Streamlit UI
col1, col2, col3, col4 = st.columns([4, 1, 1, 1])
with col1:
    st.title("💰 BetConstruct Çekim Talepleri Yönetimi")
with col2:
    # Seçilen talep bilgisi
    if 'selected_request_for_action' in st.session_state:
        request = st.session_state.selected_request_for_action
        st.info(f"Seçili: {request.get('ClientLogin', 'N/A')} - {request.get('Amount', 0)} TL")
with col2:
    pass
with col3:
    pass
with col4:
    if st.button("⚙️", help="Token ve API Ayarları"):
        st.session_state.show_settings = not st.session_state.get('show_settings', False)

# Token Ayarları Modalı
if st.session_state.get('show_settings', False):
    with st.container():
        st.markdown("### ⚙️ API Ayarları")
        
        # Mevcut ayarları göster
        col_token, col_url = st.columns(2)
        
        with col_token:
            current_token = TOKEN
            masked_token = current_token[:10] + "..." + current_token[-10:] if len(current_token) > 20 else current_token
            st.text_input("Mevcut Token", value=masked_token, disabled=True)
            
        with col_url:
            st.text_input("API URL", value=API_URL, disabled=True)
        
        # Yeni token girişi
        new_token = st.text_input("Yeni Token", placeholder="Yeni API token'ı girin...")
        new_url = st.text_input("Yeni API URL", value=API_URL, placeholder="API endpoint URL")
        
        col_save, col_cancel = st.columns(2)
        with col_save:
            if st.button("💾 Kaydet", type="primary"):
                if new_token.strip():
                    config['token'] = new_token.strip()
                if new_url.strip():
                    config['api_url'] = new_url.strip()
                
                if save_config(config):
                    # Global değişkenleri güncelle
                    update_global_config()
                    st.success("✅ Ayarlar kaydedildi! Sayfa yenileniyor...")
                    st.session_state.show_settings = False
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error("❌ Ayarlar kaydedilemedi!")
        
        with col_cancel:
            if st.button("❌ İptal"):
                st.session_state.show_settings = False
                st.rerun()

# Tab sistemi ekle
tab1, tab2, tab3, tab4 = st.tabs(["💰 Çekim Talepleri", "📊 Bahis Raporu", "📈 Performans Analizi", "🏆 Bonus Raporu"])

with tab1:
    st.markdown("---")

    # Sidebar - Filtreler (sadece Çekim Talepleri sekmesi için)
    st.sidebar.header("🔍 Filtreler")

    # Tarih seçimi
    col1, col2 = st.sidebar.columns(2)
    with col1:
        date_from = st.date_input(
            "Başlangıç Tarihi",
            value=datetime.now().date()
        )
    with col2:
        date_to = st.date_input(
            "Bitiş Tarihi", 
            value=datetime.now().date()
        )

    # Durum filtresi
    selected_status = st.sidebar.selectbox(
        "Durum Filtresi",
        options=list(STATUS_MAPPING.keys()),
        index=0
    )

    # Ödeme sistemi filtresi - dinamik olarak oluşturulacak
    if 'withdrawal_data' in st.session_state and st.session_state.withdrawal_data:
        payment_systems = set()
        for req in st.session_state.withdrawal_data:
            if isinstance(req, dict) and req.get("PaymentSystemName"):
                payment_systems.add(req["PaymentSystemName"])
        payment_options = ["Tümü"] + sorted(list(payment_systems))
    else:
        payment_options = ["Tümü", "BankTransferBME", "HedefHavale", "ScashMoneyBankTransfer"]

    payment_system_filter = st.sidebar.selectbox(
        "Ödeme Sistemi Filtresi",
        options=payment_options,
        index=0
    )

    # Sayfa boyutu
    page_size = st.sidebar.selectbox(
        "Sayfa Boyutu",
        options=[10, 25, 50, 100, 500, 1000],
        index=3  # Default 100
    )

    # İstemci tarafı filtreleme açık/kapalı
    client_side_filter = st.sidebar.checkbox(
        "🔧 İstemci Tarafı Tarih Filtrelemesi",
        value=True,
        help="API tarih filtrelemesi çalışmıyorsa otomatik olarak istemci tarafında filtreleme yapar"
    )

    # Talepleri getir butonu
    if st.sidebar.button("📋 Talepleri Gör", type="primary"):
        if date_from <= date_to:
            with st.spinner("Çekim talepleri getiriliyor..."):
                # Tüm veriyi getir, filtreleme client-side yapılacak
                all_data = []
                current_page = 1
                total_fetched = 0
                
                # Tüm verileri sayfalayarak getir
                while total_fetched < page_size:
                    remaining = page_size - total_fetched
                    current_page_size = min(100, remaining)  # API max 100'er getir
                    
                    data = fetch_withdrawal_requests(
                        datetime.combine(date_from, datetime.min.time()),
                        datetime.combine(date_to, datetime.max.time()),
                        [],
                        page=current_page,
                        page_size=current_page_size,
                        timezone_format="UTC"
                    )
                    
                    if data and isinstance(data, dict) and "Data" in data:
                        data_section = data["Data"]
                        if "ClientRequests" in data_section and data_section["ClientRequests"]:
                            page_requests = data_section["ClientRequests"]
                            all_data.extend(page_requests)
                            total_fetched += len(page_requests)
                            
                            # Eğer daha az veri geldi ise son sayfa
                            if len(page_requests) < current_page_size:
                                break
                        else:
                            break
                    else:
                        break
                    
                    current_page += 1
                    
                    # Güvenlik için max 10 sayfa
                    if current_page > 10:
                        break
                
                if all_data:
                    # İstemci tarafında tarih filtrelemesi uygula
                    if client_side_filter:
                        filtered_data = filter_requests_by_date(all_data, date_from, date_to)
                        all_data = filtered_data
                    
                    if all_data:
                        # Yeni talepler en üstte olacak şekilde sırala
                        sorted_data = sort_requests_by_status_and_date(all_data)
                        st.session_state.withdrawal_data = sorted_data
                        st.session_state.selected_status = selected_status
                        st.session_state.payment_system_filter = payment_system_filter
                        
                        # Yeni talep sayısını göster ve bildirim gönder
                        new_count = len([req for req in sorted_data 
                                       if isinstance(req, dict) and req.get("StateName") in ["Yeni", "New"]])
                        
                        # Önceki yeni talep sayısı ile karşılaştır
                        previous_new_count = st.session_state.get('previous_new_count', 0)
                        
                        # Yeni talep geldi mi kontrol et (sessizce)
                        if new_count > previous_new_count:
                            new_diff = new_count - previous_new_count
                            
                            # Basit bildirim (otomatik yenileme kaldırıldığından sadece bilgi)
                            st.toast(f"🆕 {new_diff} yeni çekim talebi tespit edildi!", icon="🔔")

                        
                        # Önceki sayıyı güncelle
                        st.session_state.previous_new_count = new_count
                    else:
                        st.warning("⚠️ Seçilen tarih aralığında veri bulunamadı")
                else:
                    st.error("❌ API'den veri alınamadı")
        else:
            st.sidebar.error("❌ Başlangıç tarihi bitiş tarihinden büyük olamaz!")

    else:
        # Auto-refresh kapalıysa timer'ı temizle
        st.markdown("""
        <script>
        if (window.autoRefreshTimer) {
            clearInterval(window.autoRefreshTimer);
            window.autoRefreshTimer = null;
            console.log('Otomatik yenileme durduruldu');
        }
        </script>
        """, unsafe_allow_html=True)

    # Ana içerik - sadece tab1 için çekim talepleri göster
    if 'withdrawal_data' in st.session_state and st.session_state.withdrawal_data:
        # Durum ve ödeme sistemi filtrelemesi uygula
        current_status_filter = st.session_state.get('selected_status', 'Tümü')
        current_payment_filter = st.session_state.get('payment_system_filter', 'Tümü')
        filtered_data = []
        
        for request in st.session_state.withdrawal_data:
            if isinstance(request, dict):
                status_text = request.get("StateName", "")
                payment_system = request.get("PaymentSystemName", "")
                
                # Durum filtresi kontrolü
                status_match = False
                if current_status_filter == "Tümü":
                    status_match = True
                else:
                    target_statuses = STATUS_MAPPING.get(current_status_filter, [])
                    status_match = status_text in target_statuses
                
                # Ödeme sistemi filtresi kontrolü
                payment_match = False
                if current_payment_filter == "Tümü":
                    payment_match = True
                else:
                    payment_match = payment_system == current_payment_filter
                
                # Her iki filtre de geçerse ekle
                if status_match and payment_match:
                    filtered_data.append(request)
        
        # Sağ üst köşeye sabit onay/red butonları (placeholder div)
        st.markdown('<div id="action-buttons-placeholder"></div>', unsafe_allow_html=True)

        # Kompakt başlık
        new_count = len([req for req in filtered_data 
                       if isinstance(req, dict) and req.get("StateName") in ["Yeni", "New"]])
        
        col1, col2, col3 = st.columns([2, 1, 1])
        with col1:
            st.markdown(f"**📊 Çekim Talepleri** ({len(filtered_data)} adet)")
        with col2:
            if new_count > 0:
                st.markdown(f"🆕 **{new_count} Yeni**")
        with col3:
            filter_text = current_status_filter
            if current_payment_filter != "Tümü":
                filter_text += f" + {current_payment_filter}"
            st.markdown(f"*Filtre: {filter_text}*")
        
        # Tablo için veri hazırlama
        table_data = []
        for idx, request in enumerate(filtered_data):
            if isinstance(request, dict):
                # Durum bilgisini al - API'den gelen StateName alanını kullan
                status_text = request.get("StateName", "Bilinmeyen")
                
                # Eğer StateName yoksa State kodunu kontrol et
                if not status_text or status_text == "Bilinmeyen":
                    state_code = request.get("State")
                    if state_code is not None:
                        status_mapping = {
                            0: "Yeni",
                            1: "İzin Verildi", 
                            2: "Beklemede",
                            3: "Reddedildi",
                            4: "İptal edildi",
                            5: "Ödendi"
                        }
                        status_text = status_mapping.get(state_code, "Bilinmeyen")
                
                # İsim bilgisi
                full_name = ""
                if "ClientName" in request:
                    full_name = str(request["ClientName"])
                elif "FirstName" in request and "LastName" in request:
                    full_name = f"{request['FirstName']} {request['LastName']}".strip()
                
                # Kullanıcı adı - API'den ClientLogin alanını kullan
                username = request.get("ClientLogin", "-")
                
                # Oyuncu kimliği - API'den ClientId alanını kullan
                player_id = str(request.get("ClientId", "-"))
                
                # Durum ile emoji gösterimi
                if status_text in ["Ödendi", "Paid"]:
                    status_display = "🟢 " + status_text
                elif status_text in ["Reddedildi", "Rejected", "Cancelled"]:
                    status_display = "🔴 " + status_text  
                elif status_text in ["Yeni", "New"]:
                    status_display = "🔵 " + status_text
                else:
                    status_display = "⚪ " + status_text
                
                table_data.append({
                    "Seç": False,
                    "Index": idx,
                    "Durum": status_display,
                    "Ödeme Sistemi": request.get("PaymentSystemName", "-"),
                    "Müşteri Adı": full_name if full_name else "-",
                    "Kullanıcı Adı": username,
                    "Oyuncu ID": player_id,
                    "Miktar (TL)": f"{float(request.get('Amount', 0)):.2f}",
                    "Talep Tarihi": request.get('RequestTimeLocal', '-')[:16] if request.get('RequestTimeLocal') else '-',
                    "Bilgi": str(request.get("Info", "-"))[:30] + "..." if len(str(request.get("Info", ""))) > 30 else str(request.get("Info", "-"))
                })
        
        if table_data:
            # DataFrame oluştur
            df = pd.DataFrame(table_data)
            
            # Veri editörü ile tablo göster
            edited_df = st.data_editor(
                df,
                column_config={
                    "Seç": st.column_config.CheckboxColumn(
                        "Seç",
                        help="Rapor için talepleri seçin",
                        default=False,
                    ),
                    "Index": None,  # Gizle
                    "Durum": st.column_config.TextColumn(
                        "Durum",
                        width="small"
                    ),
                    "Ödeme Sistemi": st.column_config.TextColumn(
                        "Ödeme Sistemi",
                        width="medium"
                    ),
                    "Miktar (TL)": st.column_config.NumberColumn(
                        "Miktar (TL)",
                        format="%.2f"
                    ),
                    "Talep Tarihi": st.column_config.TextColumn(
                        "Talep Tarihi",
                        width="medium"
                    ),
                    "Bilgi": st.column_config.TextColumn(
                        "Bilgi",
                        width="large"
                    )
                },
                hide_index=True,
                use_container_width=True
            )
        



            # Tablo seçimini kontrol et - sidebar'da işlem yapmak için
            selected_for_action = edited_df[edited_df["Seç"] == True]
            if len(selected_for_action) == 1:
                # Tek bir talep seçiliyse sidebar'da işlem için hazırla
                selected_idx = selected_for_action["Index"].iloc[0]
                st.session_state.selected_request_for_action = filtered_data[selected_idx]
            elif len(selected_for_action) > 1:
                # Birden fazla seçim varsa sidebar seçimini temizle
                if 'selected_request_for_action' in st.session_state:
                    del st.session_state.selected_request_for_action

            # Seçilen talepler için rapor oluşturma
            selected_indices = edited_df[edited_df["Seç"]]["Index"].tolist()
            
            if selected_indices:
                st.markdown("---")
                st.subheader(f"✅ {len(selected_indices)} talep seçildi")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    if st.button("📄 Çekim Raporu Oluştur", type="primary"):
                        selected_requests = [filtered_data[i] for i in selected_indices]
                        
                        # BankTransferBME taleplerini filtrele
                        bank_requests = [req for req in selected_requests if req.get("PaymentSystemName") == "BankTransferBME"]
                        
                        if bank_requests:
                            report = create_withdrawal_report(bank_requests)
                            if report:
                                st.success("✅ Çekim raporu hazırlandı!")
                                
                                # JavaScript tabanlı kopyalama butonu ile raporu göster
                                st.markdown("### 📄 Çekim Raporu")
                                create_copy_button(report, "📋 Çekim Raporunu Kopyala", "withdrawal_report")
                            else:
                                st.warning("⚠️ Rapor oluşturulamadı - veri eksik olabilir")
                        else:
                            st.warning("⚠️ Seçilen talepler arasında BankTransferBME ödeme sistemi bulunamadı")
            
            with col2:
                if st.button("🚨 Fraud Raporu Oluştur", type="secondary"):
                    if len(selected_indices) == 1:
                        # Tek bir talep seçildiyse fraud raporu oluştur
                        selected_request = filtered_data[selected_indices[0]]
                        client_id = selected_request.get("ClientId")
                        
                        if client_id:
                            with st.spinner("Fraud raporu hazırlanıyor..."):
                                fraud_report = create_fraud_report(selected_request, client_id)
                                
                                if fraud_report:
                                    st.success("✅ Fraud raporu hazırlandı!")
                                    
                                    # JavaScript tabanlı kopyalama butonu ile raporu göster
                                    st.markdown("### 🚨 Fraud Raporu")
                                    create_copy_button(fraud_report, "📋 Fraud Raporunu Kopyala", "fraud_report")
                                else:
                                    st.error("❌ Fraud raporu oluşturulamadı")
                        else:
                            st.error("❌ Client ID bulunamadı")
                    else:
                        st.warning("⚠️ Fraud raporu için sadece 1 talep seçin")
            
            # İstatistikler
            st.markdown("---")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                total_amount = sum(float(req.get("Amount", 0)) for req in filtered_data if isinstance(req, dict))
                st.metric("💰 Toplam Miktar", f"{total_amount:.2f}")
            
            with col2:
                bank_transfer_count = len([req for req in filtered_data 
                                         if isinstance(req, dict) and req.get("PaymentSystemName") == "BankTransferBME"])
                st.metric("🏦 BankTransferBME", bank_transfer_count)
            
            with col3:
                unique_users = len(set(req.get("ClientName", req.get("ClientLogin", "")) for req in filtered_data if isinstance(req, dict)))
                st.metric("👥 Benzersiz Kullanıcı", unique_users)
            
            with col4:
                st.metric("📋 Filtrelenen Talep", len(filtered_data))
        else:
            st.warning("⚠️ Seçilen filtrelere uygun talep bulunamadı")
    else:
        # Başlangıç durumu
        st.info("👋 Çekim taleplerini görüntülemek için sol panelden filtreleri ayarlayın ve 'Talepleri Gör' butonuna basın.")
        
        # Yardım bilgileri
        with st.expander("ℹ️ Kullanım Kılavuzu"):
            st.markdown("""
        ### 📋 Nasıl Kullanılır:
        1. **Tarih Aralığı Seçin**: Sol panelden başlangıç ve bitiş tarihlerini belirleyin
        2. **Durum Filtresi**: İstediğiniz duruma göre talepleri filtreleyin
        3. **Talepleri Getir**: Filtrelere uygun talepleri getirmek için butona basın
        4. **Talep Seçimi**: Tabloda rapor oluşturmak istediğiniz talepleri seçin
        5. **Rapor Oluştur**: Seçilen BankTransferBME talepleri için rapor oluşturun
        
        ### 🏦 Rapor Özellikleri:
        - Sadece **BankTransferBME** ödeme sistemi olan talepler raporda yer alır
        - Rapor otomatik olarak clipboard'a kopyalanır
        - İsim, IBAN, Banka ve Miktar bilgileri düzenli formatta hazırlanır
        
        ### 🔧 Timezone Test Özellikleri:
        - **UTC (Varsayılan)**: Standart UTC formatı
        - **UTC+3 (TR Kış)**: Türkiye kış saati formatı
        - **UTC+4 (TR Yaz)**: Türkiye yaz saati formatı  
        - **UTC Z Formatı**: ISO 8601 Z formatı
        - **Türkiye Yerel**: Otomatik timezone tespit
        
        ### 🔍 Debug Paneli:
        - API tarih filtrelemesi çalışmıyorsa otomatik istemci filtrelemesi
        - Gönderilen tarih formatı ve API yanıtı analizi
        - Farklı timezone formatlarını karşılaştırma
        """)

with tab2:
    # Sidebar'ı temizle - bu sekme için geçerli değil
    st.sidebar.empty()
    st.sidebar.header("📊 Bahis Raporu")
    st.sidebar.markdown("Bu sekmede bahis raporları görüntülenir.")
    
    st.markdown("---")
    st.header("📊 Müşteri Bahis Raporu")
    
    # Client ID girişi
    col1, col2 = st.columns([2, 1])
    with col1:
        client_id_input = st.text_input("👤 Müşteri ID'si Girin:", 
                                       placeholder="Örn: 12345", 
                                       help="Bahis raporunu görüntülemek istediğiniz müşteri ID'sini girin")
    with col2:
        # Tarih aralığı seçimi
        days_range = st.selectbox(
            "📅 Rapor Aralığı",
            options=[7, 15, 30, 60, 90],
            index=2,
            help="Kaç günlük bahis geçmişi görüntülensin?"
        )
    
    if client_id_input and client_id_input.isdigit():
        client_id = int(client_id_input)
        
        # Tarih aralığını hesapla
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days_range)
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("📈 Bahis Raporu Oluştur", type="primary"):
                with st.spinner("Bahis verileri getiriliyor..."):
                    # API'lerden veri çek
                    client_transactions = fetch_client_transactions(client_id, start_date, end_date)
                    sports_bets = fetch_client_sports_bets(client_id, start_date, end_date)
                    pending_bets = fetch_pending_sports_bets(client_id)
                    
                    if client_transactions or sports_bets or pending_bets:
                        # Rapor oluştur
                        betting_report = create_betting_report(
                            client_id, 
                            client_transactions, 
                            sports_bets, 
                            pending_bets
                        )
                        
                        if betting_report:
                            st.success("✅ Bahis raporu başarıyla oluşturuldu!")
                            
                            # Raporu göster
                            st.text_area(
                                "📋 Bahis Raporu",
                                value=betting_report,
                                height=400,
                                help="Bu raporu kopyalayabilirsiniz"
                            )
                            
                            # İndir butonu
                            st.download_button(
                                label="💾 Raporu İndir",
                                data=betting_report,
                                file_name=f"bahis_raporu_client_{client_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.txt",
                                mime="text/plain"
                            )
                        else:
                            st.error("❌ Bahis raporu oluşturulamadı")
                    else:
                        st.warning("⚠️ Bu müşteri için bahis verisi bulunamadı")
        
        with col2:
            if st.button("🔍 Müşteri Detayları"):
                with st.spinner("Müşteri bilgileri getiriliyor..."):
                    client_details = fetch_client_details(client_id)
                    client_kpis = fetch_client_kpis(client_id)
                    
                    if client_details:
                        st.subheader("👤 Müşteri Bilgileri")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric("💰 Mevcut Bakiye", f"{float(client_details.get('Balance', 0)):.2f} TL")
                            st.metric("📧 Email", client_details.get('Email', 'N/A'))
                        with col2:
                            st.metric("👤 Kullanıcı Adı", client_details.get('UserName', 'N/A'))
                            st.metric("📱 Telefon", client_details.get('MobileNumber', 'N/A'))
                        
                        # KPI bilgileri
                        if client_kpis and len(client_kpis) > 0:
                            kpi_data = client_kpis[0]
                            st.subheader("📊 KPI Bilgileri")
                            
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("💰 Toplam Yatırım", f"{float(kpi_data.get('TotalDeposit', 0)):.2f} TL")
                            with col2:
                                st.metric("💸 Toplam Çekim", f"{float(kpi_data.get('TotalWithdrawal', 0)):.2f} TL")
                            with col3:
                                st.metric("🔢 Çekim Sayısı", int(kpi_data.get('WithdrawalCount', 0)))
                    else:
                        st.error("❌ Müşteri bilgileri alınamadı")
        
        # Hızlı API testi bölümü
        with st.expander("🔧 API Test Paneli"):
            st.markdown("### API Endpoints Test")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if st.button("Test Transactions API"):
                    with st.spinner("Transactions API test ediliyor..."):
                        result = fetch_client_transactions(client_id, start_date, end_date, page_size=5)
                        if result:
                            st.json(result)
                        else:
                            st.error("Transactions API yanıt vermedi")
            
            with col2:
                if st.button("Test Sports Bets API"):
                    with st.spinner("Sports Bets API test ediliyor..."):
                        result = fetch_client_sports_bets(client_id, start_date, end_date)
                        if result:
                            st.json(result)
                        else:
                            st.error("Sports Bets API yanıt vermedi")
            
            with col3:
                if st.button("Test Pending Bets API"):
                    with st.spinner("Pending Bets API test ediliyor..."):
                        result = fetch_pending_sports_bets(client_id)
                        if result:
                            st.json(result)
                        else:
                            st.error("Pending Bets API yanıt vermedi")
    
    elif client_id_input and not client_id_input.isdigit():
        st.error("❌ Lütfen geçerli bir sayısal müşteri ID'si girin")
    
    else:
        st.info("👋 Bahis raporu oluşturmak için müşteri ID'sini girin ve 'Bahis Raporu Oluştur' butonuna basın.")
        
        # Yardım bilgileri
        with st.expander("ℹ️ Bahis Raporu Hakkında"):
            st.markdown("""
            ### 📊 Bahis Raporu İçeriği:
            1. **Genel Bahis Geçmişi**: Oyun adı, bahis miktarı ve kazanç bilgileri
            2. **En Yüksek Bahis**: Tek seferde yapılan en yüksek bahis miktarı
            3. **En Yüksek Kazanç**: Tek seferde elde edilen en yüksek kazanç
            4. **Bekleyen Spor Bahisleri**: Henüz sonuçlanmamış spor bahisleri
            
            ### 🔧 Kullanım:
            - Müşteri ID'sini girin (sayısal değer)
            - Rapor aralığını seçin (7-90 gün arası)
            - "Bahis Raporu Oluştur" butonuna basın
            - Raporu görüntüleyin ve indirin
            
            ### 📈 API Endpoints:
            - **GetClientTransactionsV1**: Genel işlem geçmişi
            - **GetClientSportBets**: Spor bahisleri
            - **GetClientPendingSportsBets**: Bekleyen bahisler
            """)

# Footer
st.markdown("---")
st.markdown("*BetConstruct Çekim Talepleri Yönetim Sistemi*")




# ===== BONUS RAPORU FONKSİYONLARI =====

def format_currency_bonus(amount):
    """Para birimi formatla (Türk Lirası) - Bonus raporu için"""
    try:
        if pd.isna(amount) or amount == "" or amount is None:
            return "0,00 TL"
        
        # Sayıya çevir
        if isinstance(amount, str):
            amount = float(amount.replace(',', '.'))
        
        # Formatla
        return f"{amount:,.2f} TL".replace(',', 'X').replace('.', ',').replace('X', '.')
    except:
        return '-'

def format_date_for_api_bonus(date_obj):
    """Tarihi BetConstruct API için formatla (dd-mm-yy - HH:MM:SS)"""
    try:
        if isinstance(date_obj, str):
            date_obj = datetime.strptime(date_obj, '%Y-%m-%d').date()
        
        if hasattr(date_obj, 'year'):
            dt = datetime.combine(date_obj, datetime.min.time())
        else:
            dt = date_obj
        
        return dt.strftime("%d-%m-%y - %H:%M:%S")
    except:
        return '-'

class BonusAPIHandler:
    def __init__(self, auth_key=None):
        self.base_url = "https://backofficewebadmin.betconstruct.com/api/tr/Report/GetClientBonusReport"
        self.auth_key = auth_key or TOKEN
        self.referer = "https://backoffice.betconstruct.com/"
        self.origin = "https://backoffice.betconstruct.com"
    
    def get_headers(self):
        """API istekleri için header oluştur"""
        return {
            "Authentication": self.auth_key,
            "Accept": "application/json",
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
            "Referer": self.referer,
            "Origin": self.origin,
            "X-Requested-With": "XMLHttpRequest"
        }
    
    def build_request_payload(self, filters):
        """API isteği için payload oluştur"""
        try:
            start_date_obj = filters["start_date"]
            end_date_obj = filters["end_date"]
            
            if isinstance(start_date_obj, str):
                start_date_obj = datetime.strptime(start_date_obj, '%Y-%m-%d').date()
            start_dt = datetime.combine(start_date_obj, datetime.min.time())
            
            if isinstance(end_date_obj, str):
                end_date_obj = datetime.strptime(end_date_obj, '%Y-%m-%d').date()
            end_dt = datetime.combine(end_date_obj, datetime.max.time())
            
            start_date = start_dt.strftime("%d-%m-%y - %H:%M:%S")
            end_date = end_dt.strftime("%d-%m-%y - %H:%M:%S")
            
            payload = {
                "ClientBonusId": "",
                "ClientId": str(filters.get("client_id", "")),
                "PartnerBonusId": "",
                "AcceptanceType": None,
                "BonusType": None,
                "BonusSource": None,
                "ByPassTotals": False,
                "EndDateLocal": None,
                "IsTest": None,
                "MaxRows": filters.get("max_rows", 100),
                "PartnerBonusEndDateLocal": end_date,
                "PartnerBonusStartDateLocal": start_date,
                "ResultFromDateLocal": None,
                "ResultToDateLocal": None,
                "ResultType": None,
                "SkeepRows": 0,
                "SportsbookProfileId": None,
                "StartDateLocal": None,
                "ToCurrencyId": "TRY"
            }
            
            return payload
            
        except Exception as e:
            return {"error": f"Payload oluşturma hatası: {str(e)}"}
    
    def get_bonus_status(self, acceptance_type):
        """Bonus durumunu çevir"""
        status_map = {
            0: "Beklemede",
            1: "Onaylandı", 
            2: "Reddedildi",
            3: "İptal Edildi"
        }
        return status_map.get(acceptance_type, "Bilinmeyen")
    
    def fetch_bonus_report(self, filters):
        """BetConstruct API'den bonus raporu getir"""
        try:
            payload = self.build_request_payload(filters)
            
            if "error" in payload:
                return {
                    "success": False,
                    "error": payload["error"],
                    "data": pd.DataFrame()
                }
            
            headers = self.get_headers()
            
            with st.spinner("Bonus raporu API'den alınıyor..."):
                response = requests.post(
                    self.base_url,
                    json=payload,
                    headers=headers,
                    timeout=30
                )
            
            if response.status_code == 200:
                data = response.json()
                df = self.process_api_response(data, filters.get("bonus_types"))
                return {
                    "success": True,
                    "data": df,
                    "total_records": len(df)
                }
            else:
                return {
                    "success": False,
                    "error": f"HTTP Hatası: {response.status_code}",
                    "data": pd.DataFrame()
                }
            
        except Exception as e:
            return {
                "success": False,
                "error": f"API hatası: {str(e)}",
                "data": pd.DataFrame()
            }
    
    def process_api_response(self, api_data, bonus_types_filter=None):
        """API yanıtını DataFrame formatına çevir"""
        try:
            if isinstance(api_data, dict) and api_data.get('HasError', False):
                return pd.DataFrame()
            
            bonus_list = None
            
            if isinstance(api_data, dict) and "Data" in api_data:
                data_obj = api_data["Data"]
                if isinstance(data_obj, dict) and "ClientBonusReportData" in data_obj:
                    bonus_report_data = data_obj["ClientBonusReportData"] 
                    if isinstance(bonus_report_data, dict) and "Objects" in bonus_report_data:
                        bonus_list = bonus_report_data["Objects"]
            
            if not bonus_list:
                return pd.DataFrame()
            
            processed_data = []
            
            for bonus in bonus_list:
                if isinstance(bonus, dict):
                    bonus_name = str(bonus.get("Name", ""))
                    
                    # Bonus türü filtresi uygula
                    if bonus_types_filter and len(bonus_types_filter) > 0:
                        # Seçilen bonus türlerinden herhangi biri ile eşleşiyor mu kontrol et
                        bonus_match = False
                        for selected_type in bonus_types_filter:
                            if bonus_name.strip().upper() == selected_type.upper():
                                bonus_match = True
                                break
                        
                        # Eşleşme yoksa bu bonusu atla
                        if not bonus_match:
                            continue
                    
                    processed_data.append({
                        'Kullanıcı ID': str(bonus.get("ClientId", "")),
                        'Kullanıcı Adı': str(bonus.get("ClientName", "")),
                        'Bonus Türü': bonus_name,
                        'Miktar': float(bonus.get("Amount", 0)),
                        'Para Birimi': str(bonus.get("ClientCurrency", "TRY")),
                        'Durum': self.get_bonus_status(bonus.get("AcceptanceType", 0)),
                        'Tarih': str(bonus.get("AcceptanceDateLocal", ""))
                    })
            
            return pd.DataFrame(processed_data)
            
        except Exception as e:
            return pd.DataFrame()

@st.cache_data(ttl=600)
def fetch_bonus_data_cached(auth_key, start_date, end_date, client_id, max_rows, bonus_types=None):
    """Önbellekli bonus veri çekme"""
    bonus_handler = BonusAPIHandler(auth_key)
    filters = {
        "start_date": start_date,
        "end_date": end_date,
        "client_id": client_id if client_id else None,
        "max_rows": max_rows,
        "bonus_types": bonus_types if bonus_types else None
    }
    return bonus_handler.fetch_bonus_report(filters)

def export_bonus_to_excel(df):
    """Bonus verilerini Excel'e aktar"""
    from openpyxl.styles import PatternFill, Font, Alignment
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Bonus Raporu", index=False)
        worksheet = writer.sheets["Bonus Raporu"]
        
        # Başlık formatı - Yeşil arka plan
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Veri formatı - Açık yeşil zebra desenli
        light_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # Hizalama - Orta
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Sütun genişlikleri ve formatlar
        for col_idx, column in enumerate(worksheet.columns, 1):
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[col_letter].width = 24
            
            # Başlık formatı
            header_cell = worksheet.cell(row=1, column=col_idx)
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = center_alignment
            
            # Veri hücreleri
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = center_alignment
                if row_idx % 2 == 0:
                    cell.fill = light_fill
        
        # Filtreleme ekle
        worksheet.auto_filter.ref = worksheet.dimensions
        
    return output.getvalue()

def export_summary_to_excel(df, filename=None):
    """Özet raporları Excel'e aktarma fonksiyonu - Geliştirilmiş formatla"""
    try:
        if df.empty:
            return None, None
        
        # Dosya adı oluştur
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"bonus_ozet_raporu_{timestamp}.xlsx"
        
        # Bellek buffer oluştur
        buffer = BytesIO()
        
        # Streamlit Cloud için xlsxwriter kullan
        try:
            with pd.ExcelWriter(buffer, engine='xlsxwriter', options={'remove_timezone': True}) as writer:
                # Ana veri sayfası
                df.to_excel(writer, sheet_name='Ana Veri', index=False)
                
                # Workbook ve formatlar
                workbook = writer.book
                
                # Başlık formatı - Koyu yeşil arka plan, beyaz yazı, kalın, ortalanmış
                header_format = workbook.add_format({
                    'bold': True,
                    'text_wrap': True,
                    'valign': 'vcenter',
                    'align': 'center',
                    'fg_color': '#70AD47',
                    'font_color': 'white',
                    'border': 1,
                    'font_size': 12
                })
                
                # Veri formatı - Açık yeşil zebra desenli, ortalanmış
                data_format_even = workbook.add_format({
                    'text_wrap': True,
                    'valign': 'vcenter',
                    'align': 'center',
                    'fg_color': '#E2EFDA',
                    'border': 1,
                    'font_size': 11
                })
                
                data_format_odd = workbook.add_format({
                    'text_wrap': True,
                    'valign': 'vcenter',
                    'align': 'center',
                    'fg_color': '#FFFFFF',
                    'border': 1,
                    'font_size': 11
                })
                
                # Ana veri sayfası için stil uygula
                worksheet = writer.sheets['Ana Veri']
                
                # Tüm sütunları 24 genişliğinde ayarla
                for col_num, value in enumerate(df.columns.values):
                    worksheet.set_column(col_num, col_num, 24)
                    # Başlık yazma
                    worksheet.write(0, col_num, value, header_format)
                
                # Veri hücrelerini formatla
                for row_num in range(1, len(df) + 1):
                    for col_num in range(len(df.columns)):
                        cell_value = df.iloc[row_num-1, col_num]
                        # Zebra desen için çift/tek satır kontrolü
                        if row_num % 2 == 0:
                            worksheet.write(row_num, col_num, cell_value, data_format_even)
                        else:
                            worksheet.write(row_num, col_num, cell_value, data_format_odd)
                
                # Auto filter ekleme
                worksheet.autofilter(0, 0, len(df), len(df.columns) - 1)
                
                # Satır yüksekliği ayarla
                worksheet.set_default_row(25)
            
            buffer.seek(0)
            return buffer.getvalue(), filename
            
        except Exception as xlsxwriter_error:
            # Fallback: openpyxl ile geliştirilmiş format
            try:
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Ana Veri', index=False)
                    worksheet = writer.sheets['Ana Veri']
                    
                    # Başlık formatı - Koyu yeşil arka plan, beyaz yazı
                    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True, size=12)
                    
                    # Veri formatı - Açık yeşil zebra desenli
                    light_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    data_font = Font(size=11)
                    
                    # Kenarlık
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Orta hizalama
                    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    # Sütun genişlikleri ve formatlar
                    for col_idx, column in enumerate(worksheet.columns, 1):
                        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                        worksheet.column_dimensions[col_letter].width = 24
                        
                        # Başlık formatı
                        header_cell = worksheet.cell(row=1, column=col_idx)
                        header_cell.fill = header_fill
                        header_cell.font = header_font
                        header_cell.alignment = center_alignment
                        header_cell.border = thin_border
                        
                        # Veri hücreleri
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = center_alignment
                            cell.font = data_font
                            cell.border = thin_border
                            # Zebra desen
                            if row_idx % 2 == 0:
                                cell.fill = light_fill
                            else:
                                cell.fill = white_fill
                    
                    # Satır yüksekliği
                    for row in range(1, worksheet.max_row + 1):
                        worksheet.row_dimensions[row].height = 25
                    
                    # Filtreleme ekle
                    worksheet.auto_filter.ref = worksheet.dimensions
                    
                buffer.seek(0)
                return buffer.getvalue(), filename
            except Exception as openpyxl_error:
                return None, None
        
    except Exception as e:
        return None, None

def create_bonus_summary_report(df):
    """Kullanıcı bazlı bonus özet raporu oluştur"""
    try:
        if df.empty:
            return pd.DataFrame()
        
        if 'Kullanıcı ID' not in df.columns or 'Bonus Türü' not in df.columns or 'Miktar' not in df.columns:
            return pd.DataFrame()
        
        # Kullanıcı ve bonus türü bazlı gruplama
        user_bonus_summary = df.groupby(['Kullanıcı ID', 'Kullanıcı Adı', 'Bonus Türü']).agg({
            'Miktar': ['count', 'sum']
        }).reset_index()
        
        # MultiIndex sütunları düzelt
        user_bonus_summary.columns = ['Kullanıcı ID', 'Kullanıcı Adı', 'Bonus Türü', 'Kaç Defa Aldı', 'Toplam Miktar']
        
        # Miktarları formatla
        user_bonus_summary['Toplam Miktar Formatted'] = user_bonus_summary['Toplam Miktar'].apply(lambda x: format_currency_bonus(x))
        
        # Sıralama
        user_bonus_summary = user_bonus_summary.sort_values(['Toplam Miktar', 'Kullanıcı ID'], ascending=[False, True])
        
        # Görüntüleme için sütun sırası
        display_columns = ['Kullanıcı ID', 'Kullanıcı Adı', 'Bonus Türü', 'Kaç Defa Aldı', 'Toplam Miktar Formatted']
        user_bonus_summary_display = user_bonus_summary[display_columns].copy()
        user_bonus_summary_display.columns = ['Kullanıcı ID', 'Kullanıcı Adı', 'Bonus Türü', 'Kaç Defa Aldı', 'Toplam Miktar']
        
        return user_bonus_summary_display
    
    except Exception as e:
        return pd.DataFrame()

def format_currency_bonus(amount):
    """Bonus miktar formatı"""
    try:
        return f"{amount:,.2f} TL"
    except:
        return "0.00 TL"

def create_bonus_type_summary(df):
    """Bonus türü bazlı özet rapor oluştur"""
    try:
        if df.empty:
            return pd.DataFrame()
        
        if 'Bonus Türü' not in df.columns:
            return pd.DataFrame()
        
        summary = df.groupby('Bonus Türü').agg({
            'Kullanıcı ID': 'count',
            'Miktar': ['sum', 'mean']
        }).reset_index()
        
        # MultiIndex sütunları düzelt
        summary.columns = ['Bonus Türü', 'Adet', 'Toplam Miktar', 'Ortalama Miktar']
        
        # Formatla
        summary['Toplam Miktar'] = summary['Toplam Miktar'].apply(lambda x: format_currency_bonus(x))
        summary['Ortalama Miktar'] = summary['Ortalama Miktar'].apply(lambda x: format_currency_bonus(x))
        
        return summary
    
    except Exception as e:
        return pd.DataFrame()

# ===== PERFORMANS ANALİZİ FONKSİYONLARI =====

def get_status_display_performance(state, allow_user, reject_user):
    """Durum gösterimi için güvenilir fonksiyon"""
    if pd.isna(state):
        return "❓ Bilinmiyor"
    state = int(state)

    # AllowUserName dolu → Ödendi
    if pd.notna(allow_user) and str(allow_user).strip() not in ["", "None", "null"]:
        return "✅ Ödendi"

    # RejectUserName dolu → Reddedildi  
    if pd.notna(reject_user) and str(reject_user).strip() not in ["", "None", "null"]:
        return "❌ Reddedildi"

    # State göre
    if state == 5:
        return "🔄 İptal Edildi"
    elif state == -1:
        return "🟠 İptal Beklemede"
    elif state in [2, -2]:
        return "⏳ Beklemede"
    else:
        return f"🔵 Bilinmeyen ({state})"

def process_data_for_performance(raw_data):
    """Performans analizi için veri işleme"""
    df = pd.DataFrame(raw_data)

    # Zorunlu sütunlar
    cols = [
        "Id", "State", "StateName", "PaymentSystemName", "ClientId", "ClientName",
        "Amount", "AllowUserName", "RejectUserName", "Info", "RequestTimeLocal", "AllowTimeLocal"
    ]
    for col in cols:
        if col not in df.columns:
            df[col] = None

    # Zaman formatı
    df["RequestTimeLocal"] = pd.to_datetime(df["RequestTimeLocal"], errors="coerce")
    df["AllowTimeLocal"] = pd.to_datetime(df["AllowTimeLocal"], errors="coerce")

    # Müşteri adı düzeltme: "Soyad Ad" → "Ad Soyad"
    def fix_name(name):
        if pd.isna(name) or not isinstance(name, str):
            return "Bilinmiyor"
        parts = name.strip().split()
        if len(parts) > 1:
            return " ".join(parts[1:] + [parts[0]])  # Ad Soyad
        return name.strip()

    df["ClientNameFormatted"] = df["ClientName"].apply(fix_name)

    # Onaylayan kullanıcı
    df["Personel"] = df["AllowUserName"].fillna(df["RejectUserName"])
    df["Personel"] = df["Personel"].fillna("–")

    # Bilgi kısaltma
    def extract_info(info):
        if pd.isna(info) or not isinstance(info, str):
            return "–"
        if "IBAN:" in info:
            return info.split("IBAN:")[1].split(",")[0][:26] + "..."
        elif "AccountNumber:" in info:
            return info.split("AccountNumber:")[1].split(",")[0][:26] + "..."
        elif "fullname:" in info:
            return info.split("fullname:")[1].split(",")[0][:26] + "..."
        return info[:30] + "..."

    df["InfoShort"] = df["Info"].apply(extract_info)

    # İşlem süresi (dakika)
    df["ProcessingTimeSec"] = (df["AllowTimeLocal"] - df["RequestTimeLocal"]).dt.total_seconds()
    df["ProcessingTimeMin"] = (df["ProcessingTimeSec"] / 60).round(2)
    df["ProcessingTimeMin"] = df["ProcessingTimeMin"].fillna(0)

    # Durum hesaplama
    df["StatusDisplay"] = df.apply(
        lambda row: get_status_display_performance(row["State"], row["AllowUserName"], row["RejectUserName"]),
        axis=1
    )

    # En yeni işlemler en üstte olacak şekilde sırala
    df = df.sort_values("RequestTimeLocal", ascending=False, na_position='last')

    return df

def calculate_performance(df):
    """Personel performansını hesapla"""
    # Sadece işlem görmüşler (ödenmiş veya reddedilmiş)
    processed = df.dropna(subset=["AllowUserName", "RejectUserName"], how="all")
    if processed.empty:
        return pd.DataFrame(columns=["Personel", "İşlemAdedi", "OrtalamaSüre"])

    perf = processed.groupby("Personel").agg(
        İşlemAdedi=("Id", "count"),
        OrtalamaSüre=("ProcessingTimeMin", "mean")
    ).round(2).reset_index()
    return perf

def export_to_excel_performance(main_df, perf_df):
    """Performans analizi için Excel'e aktar"""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Ana işlemler sayfası
        main_df.to_excel(writer, sheet_name="İşlemler", index=False)
        worksheet1 = writer.sheets["İşlemler"]
        
        # Başlık formatı - Mavi arka plan
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Veri formatı - Açık gri zebra desenli
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Hizalama - Orta
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Sütun genişliklerini 24 yap ve formatla
        for col_idx, column in enumerate(worksheet1.columns, 1):
            col_letter = worksheet1.cell(row=1, column=col_idx).column_letter
            worksheet1.column_dimensions[col_letter].width = 24
            
            # Başlık formatı
            header_cell = worksheet1.cell(row=1, column=col_idx)
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = center_alignment
            
            # Veri hücrelerini formatla
            for row_idx in range(2, worksheet1.max_row + 1):
                cell = worksheet1.cell(row=row_idx, column=col_idx)
                cell.alignment = center_alignment
                # Zebra desen
                if row_idx % 2 == 0:
                    cell.fill = light_fill
        
        # Filtreleme ekle
        worksheet1.auto_filter.ref = worksheet1.dimensions
        
        # Performans sayfası
        if not perf_df.empty:
            perf_df.to_excel(writer, sheet_name="Performans", index=False)
            worksheet2 = writer.sheets["Performans"]
            
            # Performans sayfası için aynı formatı uygula
            for col_idx, column in enumerate(worksheet2.columns, 1):
                col_letter = worksheet2.cell(row=1, column=col_idx).column_letter
                worksheet2.column_dimensions[col_letter].width = 24
                
                # Başlık formatı
                header_cell = worksheet2.cell(row=1, column=col_idx)
                header_cell.fill = header_fill
                header_cell.font = header_font
                header_cell.alignment = center_alignment
                
                # Veri hücrelerini formatla
                for row_idx in range(2, worksheet2.max_row + 1):
                    cell = worksheet2.cell(row=row_idx, column=col_idx)
                    cell.alignment = center_alignment
                    if row_idx % 2 == 0:
                        cell.fill = light_fill
            
            # Filtreleme ekle
            worksheet2.auto_filter.ref = worksheet2.dimensions
            
    return output.getvalue()

@st.cache_data(ttl=600)  # 10 dakika önbellek
def fetch_withdrawal_requests_for_performance(token):
    """Performans analizi için çekim taleplerini API'den getir"""
    headers = {
        "Content-Type": "application/json;charset=UTF-8",
        "Authentication": token,
        "Accept": "application/json, text/plain, */*",
        "Origin": "https://backoffice.betconstruct.com",  
        "Referer": "https://backoffice.betconstruct.com/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }
    payload = {
        "PageIndex": 1,
        "PageSize": 500,
        "Sort": {"Field": "Id", "Dir": "desc"},
        "DateFrom": None,
        "DateTo": None,
        "Statuses": [],
        "PaymentProviderId": None,
        "SearchText": "",
        "WithdrawId": None
    }

    try:
        with st.spinner("API'den veri alınıyor..."):
            response = requests.post(API_URL, json=payload, headers=headers, timeout=30)
            if response.status_code == 200:
                data = response.json()
                if data.get("HasError"):
                    st.error(f"API Hatası: {data.get('AlertMessage', 'Bilinmeyen hata')}")
                    return None
                if "Data" in data and "ClientRequests" in data["Data"]:
                    return data["Data"]["ClientRequests"]
                else:
                    st.error("API yanıtında beklenen veri yapısı bulunamadı.")
                    return None
            else:
                st.error(f"HTTP Hatası: {response.status_code} - {response.text}")
                return None
    except Exception as e:
        st.error(f"Bağlantı hatası: {str(e)}")
        return None

# TAB 3 - PERFORMANS ANALİZİ
with tab3:
    # Sidebar'ı temizle - bu sekme için geçerli değil
    st.sidebar.empty()
    st.sidebar.header("📈 Performans Analizi")
    st.sidebar.markdown("Bu sekmede performans analizleri görüntülenir.")
    
    st.header("📈 Performans Analizi")
    
    # Bugünün tarihi olarak varsayılan filtre
    today = datetime.now().date()
    
    col1, col2 = st.columns(2)
    with col1:
        date_from = st.date_input("Başlangıç Tarihi", value=today, key="perf_date_from")
    with col2:
        date_to = st.date_input("Bitiş Tarihi", value=today, key="perf_date_to")

    # Token kontrolü
    if not TOKEN.strip():
        st.warning("Lütfen ayarlardan bir API token girin.")
        st.stop()

    # Veri çek (istemci taraflı filtreleme için tüm veriler)
    raw_data = fetch_withdrawal_requests_for_performance(TOKEN.strip())
    if raw_data is None or len(raw_data) == 0:
        st.error("Veri alınamadı. Token veya ağ bağlantınızı kontrol edin.")
        st.stop()

    # Veriyi işle
    df = process_data_for_performance(raw_data)
    
    # İstemci taraflı tarih filtresi uygula
    if date_from and date_to:
        mask = (df["RequestTimeLocal"].dt.date >= date_from) & (df["RequestTimeLocal"].dt.date <= date_to)
        filtered_df = df[mask].copy()
    else:
        filtered_df = df.copy()

    # Performans hesapla
    perf_df = calculate_performance(filtered_df)

    if perf_df.empty:
        st.warning("Seçilen tarih aralığında performans verisi bulunamadı.")
    else:
        # Grafikler
        col1, col2 = st.columns(2)
        with col1:
            fig1 = px.bar(perf_df, x="Personel", y="İşlemAdedi", title="İşlem Adedi", color="İşlemAdedi")
            st.plotly_chart(fig1, use_container_width=True)
        with col2:
            fig2 = px.pie(perf_df, names="Personel", values="İşlemAdedi", title="İşlem Dağılımı")
            st.plotly_chart(fig2, use_container_width=True)

        fig3 = px.bar(perf_df, x="Personel", y="OrtalamaSüre", title="Ortalama Süre (dakika)", color="OrtalamaSüre")
        st.plotly_chart(fig3, use_container_width=True)

        # Performans tablosu
        st.subheader("Personel Performansı")
        st.dataframe(perf_df, use_container_width=True, hide_index=True)

        # Excel aktar butonu
        excel_data = export_to_excel_performance(
            filtered_df[[
                "StatusDisplay", "PaymentSystemName", "ClientNameFormatted", "ClientId",
                "Amount", "Personel", "InfoShort", "ProcessingTimeMin"
            ]].rename(columns={
                "StatusDisplay": "Durum", "PaymentSystemName": "Ödeme Sistemi", "ClientNameFormatted": "Müşteri Adı",
                "ClientId": "Oyuncu ID", "Amount": "Miktar", "Personel": "Personel",
                "InfoShort": "Bilgi", "ProcessingTimeMin": "Süre (dk)"
            }),
            perf_df
        )
        st.download_button(
            label="📥 Excel'e Aktar",
            data=excel_data,
            file_name=f"Performans_Raporu_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# TAB 4 - BONUS RAPORU
with tab4:
    # Sidebar'ı temizle - bu sekme için geçerli değil
    st.sidebar.empty()
    st.sidebar.header("🏆 Bonus Raporu")
    st.sidebar.markdown("Bu sekmede bonus raporları görüntülenir.")
    
    st.header("🏆 Bonus Raporu")
    
    # Dünün tarihi olarak varsayılan filtre
    yesterday = datetime.now().date() - timedelta(days=1)
    
    col1, col2 = st.columns(2)
    with col1:
        bonus_start_date = st.date_input("Başlangıç Tarihi", value=yesterday, key="bonus_start_date")
    with col2:
        bonus_end_date = st.date_input("Bitiş Tarihi", value=yesterday, key="bonus_end_date")

    # Bonus türü filtresi
    bonus_types = [
        "CASİNO KAYIP BONUSU", "%100 SLOT BONUSU", "%100 CASİNO HOŞGELDİN BONUSU",
        "%100 PRAGMATİC SALI - PERŞEMBE", "%100 SPOR HOŞGELDİN BONUSU",
        "%25 SPOR YATIRIM BONUSU", "%5 CASİNO HAFTALIK", "%5 SPOR HAFTALIK",
        "250 TL CASİNO DENEME BONUSU", "250 TL DOĞUM GÜNÜ CASİNO BONUSU",
        "250 TL SPOR DENEME BONUSU", "CASİNO BAĞLILIK BONUSU", "CASİNO CALL DAVET",
        "CASİNO ÇEVRİMSİZ BONUS", "CASİNO DOĞUM GÜNÜ BONUSU", "%10 ÇEVRİMSİZ SPOR BONUSU",
        "P.TESİ & ÇARŞAMBA %100 GÜNÜN İLK KAYIBINA", "SPOR BAĞLILIK BONUSU",
        "SPOR CALL DAVET", "SPOR ÇEVRİMSİZ BONUS", "SPOR DOĞUM GÜNÜ BONUSU",
        "SPOR KAYIP BONUSU", "YENİ CASİNO ŞANS BONUSU"
    ]
    
    selected_bonus_types = st.multiselect(
        "Bonus Türleri (Birden fazla seçebilirsiniz):",
        bonus_types,
        help="Boş bırakırsanız tüm bonus türleri getirilir",
        key="bonus_types_filter"
    )

    # Diğer filtreler
    col3, col4 = st.columns(2)
    with col3:
        bonus_client_id = st.text_input("Kullanıcı ID (isteğe bağlı):", key="bonus_client_id")
    with col4:
        bonus_max_rows = st.number_input("Maksimum Kayıt:", min_value=1, max_value=10000, value=1000, key="bonus_max_rows")

    # Token kontrolü
    if not TOKEN.strip():
        st.warning("Lütfen ayarlardan bir API token girin.")
        st.stop()

    # Veri getir butonu
    if st.button("🔍 Bonus Raporunu Getir", type="primary"):
        if bonus_start_date > bonus_end_date:
            st.error("Başlangıç tarihi bitiş tarihinden sonra olamaz!")
        else:
            try:
                result = fetch_bonus_data_cached(
                    TOKEN.strip(),
                    bonus_start_date,
                    bonus_end_date,
                    bonus_client_id.strip() if bonus_client_id else None,
                    bonus_max_rows,
                    selected_bonus_types if selected_bonus_types else None
                )
                
                if result["success"]:
                    st.session_state.bonus_data = result["data"]
                    st.success(f"✅ {result['total_records']} bonus kaydı getirildi!")
                else:
                    st.error(f"❌ {result['error']}")
                    st.session_state.bonus_data = pd.DataFrame()
                    
            except Exception as e:
                st.error(f"❌ Beklenmeyen hata: {str(e)}")
                st.session_state.bonus_data = pd.DataFrame()

    # Sonuçları göster
    if 'bonus_data' in st.session_state and not st.session_state.bonus_data.empty:
        st.subheader("📋 Bonus Raporu Sonuçları")
        
        # Veri tablosu
        st.dataframe(st.session_state.bonus_data, use_container_width=True, height=400)
        
        # Özet bilgiler
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.info(f"📊 Toplam kayıt: {len(st.session_state.bonus_data)}")
        
        with col2:
            if 'Miktar' in st.session_state.bonus_data.columns:
                total_amount = st.session_state.bonus_data['Miktar'].sum()
                st.info(f"💰 Toplam miktar: {format_currency_bonus(total_amount)}")
        
        with col3:
            if 'Kullanıcı ID' in st.session_state.bonus_data.columns:
                unique_users = st.session_state.bonus_data['Kullanıcı ID'].nunique()
                st.info(f"👤 Benzersiz kullanıcı: {unique_users}")
        
        # Bonus türlerine göre dağılım
        if 'Bonus Türü' in st.session_state.bonus_data.columns:
            st.subheader("📈 Bonus Türü Dağılımı")
            bonus_counts = st.session_state.bonus_data['Bonus Türü'].value_counts()
            
            col1, col2 = st.columns(2)
            with col1:
                fig_bar = px.bar(
                    x=bonus_counts.index,
                    y=bonus_counts.values,
                    title="Bonus Türü Adedi",
                    labels={'x': 'Bonus Türü', 'y': 'Adet'}
                )
                fig_bar.update_xaxes(tickangle=45)
                st.plotly_chart(fig_bar, use_container_width=True)
            
            with col2:
                fig_pie = px.pie(
                    values=bonus_counts.values,
                    names=bonus_counts.index,
                    title="Bonus Türü Oranı"
                )
                st.plotly_chart(fig_pie, use_container_width=True)
        
        # İşlem butonları
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # Excel export butonu
            excel_data = export_bonus_to_excel(st.session_state.bonus_data)
            st.download_button(
                label="📥 Excel'e Aktar",
                data=excel_data,
                file_name=f"Bonus_Raporu_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        with col2:
            # Özet rapor butonu
            if st.button("📈 Özet Rapor Oluştur"):
                # Kullanıcı bazlı özet rapor
                user_summary = create_bonus_summary_report(st.session_state.bonus_data)
                
                if not user_summary.empty:
                    st.subheader("👥 Kullanıcı Bazlı Özet")
                    st.dataframe(user_summary, use_container_width=True, hide_index=True)
                    
                    # Kullanıcı özet raporu Excel indirme
                    try:
                        excel_data = export_summary_to_excel(user_summary, f"kullanici_ozet_{bonus_start_date.strftime('%Y%m%d')}_{bonus_end_date.strftime('%Y%m%d')}.xlsx")
                        if excel_data[0] and excel_data[1]:
                            st.download_button(
                                label="📥 Kullanıcı Özet Raporunu İndir",
                                data=excel_data[0],
                                file_name=excel_data[1],
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key="user_summary_download"
                            )
                        else:
                            st.error("Kullanıcı özet rapor dosyası oluşturulamadı.")
                    except Exception as e:
                        st.error(f"Kullanıcı özet rapor indirme hatası: {str(e)}")
                    
                    st.divider()
                    
                    # Bonus türü bazlı özet
                    bonus_type_summary = create_bonus_type_summary(st.session_state.bonus_data)
                    if not bonus_type_summary.empty:
                        st.subheader("🎁 Bonus Türü Bazlı Özet")
                        st.dataframe(bonus_type_summary, use_container_width=True, hide_index=True)
                        
                        # Bonus türü özet raporu Excel indirme
                        try:
                            excel_data = export_summary_to_excel(bonus_type_summary, f"bonus_turu_ozet_{bonus_start_date.strftime('%Y%m%d')}_{bonus_end_date.strftime('%Y%m%d')}.xlsx")
                            if excel_data[0] and excel_data[1]:
                                st.download_button(
                                    label="📥 Bonus Türü Özet Raporunu İndir",
                                    data=excel_data[0],
                                    file_name=excel_data[1],
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                    key="bonus_type_summary_download"
                                )
                            else:
                                st.error("Bonus türü özet rapor dosyası oluşturulamadı.")
                        except Exception as e:
                            st.error(f"Bonus türü özet rapor indirme hatası: {str(e)}")
                    
                else:
                    st.warning("Özet rapor oluşturulamadı.")
        
        with col3:
            # Temizle butonu
            if st.button("🗑️ Sonuçları Temizle"):
                st.session_state.bonus_data = pd.DataFrame()
                st.success("Sonuçlar temizlendi!")
                st.rerun()
    
    elif 'bonus_data' in st.session_state and st.session_state.bonus_data.empty:
        st.info("📝 Bonus raporu getirmek için yukarıdaki butona tıklayın.")
    else:
        # Session state'i başlat
        if 'bonus_data' not in st.session_state:
            st.session_state.bonus_data = pd.DataFrame()
        st.info("📝 Bonus raporu getirmek için yukarıdaki butona tıklayın.")

# Footer
st.markdown("---")
st.markdown("*BetConstruct Çekim Talepleri Yönetimi v2.0*")

